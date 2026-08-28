"""The management area's HTTP contract, over a fake repository and model.

Three things here are the point of the feature and are asserted directly:

- **A member cannot write.** Every mutating route depends on `guards.boss()`,
  so a member's cookie is refused whichever URL it is aimed at (D5).
- **One workspace cannot reach another's schedule**, even holding its id
  (D10). A cross-team miss reads as "not found", never as "not yours".
- **Warnings never block.** A schedule that breaks every rule still returns
  200 and still renders (D3).
"""

from types import SimpleNamespace
import json
import threading
import time

import pytest
from fastapi import FastAPI
from fastapi.responses import JSONResponse
from fastapi.testclient import TestClient

from app.api.dependencies import Guards
from app.api.routers import schedules
from app.bl.schedule_service import ScheduleService
from app.common.errors import AgentError, AppError, NotFoundError, error_payload
from app.common.sessions import COOKIE_NAME, ROLE_BOSS, ROLE_MEMBER, issue

TEAM = "team-a"
OTHER_TEAM = "team-b"
SECRET = "test-secret"

MORNING = "בוקר"
EVENING = "צהריים"

PROFILE = {
    "workplace": {"name": "מוקד", "mission": "מענה", "planning_horizon": "שבוע"},
    "employees": [
        {"name": "דנה", "eligible_shifts": [MORNING]},
        {"name": "יוסי", "eligible_shifts": [MORNING, EVENING]},
    ],
    "shifts": [{
        "name": MORNING, "start_time": "07:00", "end_time": "15:00",
        "days": [], "is_on_call": False, "hour_weight": 1.0,
        "staffing": [{"days": [], "headcount": 1, "required_roles": []}],
    }],
    "rules": [],
}


class _FakeScheduleRepo:
    """In-memory stand-in that filters by team exactly as the SQL does."""

    def __init__(self):
        self.schedules = {}
        self.slots = {}
        self.assignment_rows = {}
        self.availability_rows = []
        self.changes = []
        self.requests = []
        self.preference_rows = []
        self.profiles = {TEAM: PROFILE, OTHER_TEAM: PROFILE}
        self._n = 0

    def _id(self, prefix):
        self._n += 1
        return "%s-%d" % (prefix, self._n)

    def team_profile(self, team_id):
        return self.profiles.get(team_id)

    def update_team_profile(self, team_id, profile):
        self.profiles[team_id] = profile
        return profile

    def create_schedule(self, team_id, starts_on, ends_on):
        schedule_id = self._id("sched")
        self.schedules[schedule_id] = {
            "id": schedule_id, "team_id": team_id, "starts_on": starts_on,
            "ends_on": ends_on, "status": "draft", "generation": {},
        }
        self.slots[schedule_id] = []
        self.assignment_rows[schedule_id] = []
        return self.get_schedule(schedule_id, team_id)

    def get_schedule(self, schedule_id, team_id):
        row = self.schedules.get(schedule_id)
        # A schedule in another workspace is indistinguishable from one that
        # does not exist -- distinguishing them would turn this into an
        # oracle for which rows live in workspaces the caller cannot see.
        if row is None or row["team_id"] != team_id:
            raise NotFoundError("הפריט לא נמצא")
        view = dict(row)
        view["slots"] = self.slots.get(schedule_id, [])
        view["assignments"] = self.assignments(schedule_id, team_id)
        return view

    def list_schedules(self, team_id):
        return [
            {k: row[k] for k in ("id", "starts_on", "ends_on", "status")}
            for row in self.schedules.values() if row["team_id"] == team_id
        ]

    def current_schedule(self, team_id, published_only=False):
        for row in reversed(list(self.schedules.values())):
            if row["team_id"] != team_id:
                continue
            if published_only and row["status"] != "published":
                continue
            return self.get_schedule(row["id"], team_id)
        return None

    def set_schedule_status(self, schedule_id, team_id, status):
        self.get_schedule(schedule_id, team_id)
        self.schedules[schedule_id]["status"] = status
        return self.get_schedule(schedule_id, team_id)

    def set_generation(self, schedule_id, team_id, generation):
        self.get_schedule(schedule_id, team_id)
        self.schedules[schedule_id]["generation"] = generation
        return self.get_schedule(schedule_id, team_id)

    def touch_generation(self, schedule_id, team_id, at):
        # Scoped to a running job exactly as the UPDATE is: a beat that lands
        # after the job stopped must not make it look alive again.
        row = self.schedules.get(schedule_id)
        if row is None or row["team_id"] != team_id:
            return
        generation = row.get("generation") or {}
        if generation.get("status") == "running":
            generation["heartbeat"] = at

    def delete_schedule(self, schedule_id, team_id):
        self.get_schedule(schedule_id, team_id)
        del self.schedules[schedule_id]

    def replace_slots(self, schedule_id, team_id, slots):
        self.get_schedule(schedule_id, team_id)
        self.slots[schedule_id] = [
            dict(slot, id=self._id("slot"), team_id=team_id) for slot in slots
        ]
        return self.slots[schedule_id]

    def find_slot(self, schedule_id, team_id, shift_name, slot_date):
        self.get_schedule(schedule_id, team_id)
        for slot in self.slots.get(schedule_id, []):
            if (slot["shift_name"] == shift_name
                    and slot["slot_date"] == slot_date):
                return slot
        return None

    def assignments(self, schedule_id, team_id):
        rows = []
        for row in self.assignment_rows.get(schedule_id, []):
            slot = next(
                s for s in self.slots[schedule_id] if s["id"] == row["slot_id"]
            )
            rows.append(dict(
                row, shift=slot["shift_name"], date=slot["slot_date"],
                start_time=slot.get("start_time", ""),
                end_time=slot.get("end_time", ""),
                is_on_call=slot.get("is_on_call", False),
                source=row.get("source", "agent"),
            ))
        return rows

    def replace_assignments(self, schedule_id, team_id, assignments):
        self.get_schedule(schedule_id, team_id)
        for item in assignments:
            assert item["reason"].strip(), "reason is required (D8)"
        self.assignment_rows[schedule_id] = [
            {"id": self._id("asg"), "slot_id": item["slot_id"],
             "employee": item["employee"], "reason": item["reason"],
             "schedule_id": schedule_id,
             "source": item.get("source") or "agent"}
            for item in assignments
        ]
        return self.assignments(schedule_id, team_id)

    def replace_span_assignments(self, schedule_id, team_id, dates,
                                 assignments):
        # Rows on other dates keep their identity, exactly as the real
        # DELETE ... WHERE slot_date = ANY(...) leaves them alone. Tests that
        # hold an assignment id across a checkpoint depend on that.
        self.get_schedule(schedule_id, team_id)
        for item in assignments:
            assert item["reason"].strip(), "reason is required (D8)"
        wanted = set(dates or [])
        slots = {slot["id"]: slot for slot in self.slots.get(schedule_id, [])}
        kept = [
            row for row in self.assignment_rows.get(schedule_id, [])
            if slots.get(row["slot_id"], {}).get("slot_date") not in wanted
        ]
        self.assignment_rows[schedule_id] = kept + [
            {"id": self._id("asg"), "slot_id": item["slot_id"],
             "employee": item["employee"], "reason": item["reason"],
             "schedule_id": schedule_id,
             "source": item.get("source") or "agent"}
            for item in assignments
        ]
        return self.assignments(schedule_id, team_id)

    def add_assignment(self, schedule_id, team_id, slot_id, employee, reason,
                       source="agent"):
        assert reason.strip(), "reason is required (D8)"
        assert source in ("agent", "manager", "imported"), source
        self.get_schedule(schedule_id, team_id)
        # The real table is UNIQUE (slot_id, employee) and the insert says
        # DO NOTHING, so placing the same person on the same slot twice is a
        # success that changes nothing. Mirrored here because the manual path
        # depends on that being the behaviour rather than a duplicate row.
        for existing in self.assignment_rows[schedule_id]:
            if (existing["slot_id"] == slot_id
                    and existing["employee"] == employee):
                return dict(existing)
        row = {"id": self._id("asg"), "slot_id": slot_id,
               "employee": employee, "reason": reason,
               "schedule_id": schedule_id, "source": source}
        self.assignment_rows[schedule_id].append(row)
        return row

    def remove_assignment(self, assignment_id, team_id):
        for schedule_id, rows in self.assignment_rows.items():
            if self.schedules[schedule_id]["team_id"] != team_id:
                continue
            self.assignment_rows[schedule_id] = [
                row for row in rows if row["id"] != assignment_id
            ]

    def move_assignment(self, assignment_id, team_id, slot_id, reason,
                        employee=None):
        assert reason.strip(), "reason is required (D8)"
        for schedule_id, rows in self.assignment_rows.items():
            if self.schedules[schedule_id]["team_id"] != team_id:
                continue
            for row in rows:
                if row["id"] == assignment_id:
                    row["slot_id"] = slot_id
                    row["reason"] = reason
                    if employee:
                        row["employee"] = employee
                    return dict(row)
        raise NotFoundError("השיבוץ לא נמצא")

    def set_availability(self, team_id, employee, constraint_date,
                         shift_name="", available=False, reason="",
                         start_time="", end_time="", is_hard=True,
                         source="manager"):
        row = {
            "id": self._id("av"), "team_id": team_id, "employee": employee,
            "constraint_date": constraint_date, "shift_name": shift_name,
            "available": available, "start_time": start_time,
            "end_time": end_time, "is_hard": is_hard,
            "reason": reason, "source": source,
        }
        self.availability_rows = [
            existing for existing in self.availability_rows
            if not (existing["team_id"] == team_id
                    and existing["employee"] == employee
                    and existing["constraint_date"] == constraint_date
                    and existing["shift_name"] == shift_name)
        ]
        self.availability_rows.append(row)
        return row

    def availability(self, team_id, starts_on=None, ends_on=None,
                     employee=None):
        rows = [
            row for row in self.availability_rows
            if row["team_id"] == team_id
        ]
        if starts_on:
            rows = [r for r in rows if r["constraint_date"] >= starts_on]
        if ends_on:
            rows = [r for r in rows if r["constraint_date"] <= ends_on]
        if employee:
            rows = [r for r in rows if r["employee"] == employee]
        return rows

    def preferences(self, team_id, status=None):
        rows = [
            row for row in self.preference_rows if row["team_id"] == team_id
        ]
        return [row for row in rows if row["status"] == status] if status else rows

    def delete_availability(self, row_id, team_id):
        self.availability_rows = [
            row for row in self.availability_rows
            if not (row["id"] == row_id and row["team_id"] == team_id)
        ]

    def append_change(self, team_id, action, schedule_id=None, employee="",
                      replaced_employee="", slot_date=None, shift_name="",
                      reason="", agent_reason=""):
        row = {
            "id": self._id("chg"), "team_id": team_id, "action": action,
            "schedule_id": schedule_id, "employee": employee,
            "replaced_employee": replaced_employee, "slot_date": slot_date,
            "shift_name": shift_name, "reason": reason,
            "agent_reason": agent_reason,
        }
        self.changes.append(row)
        return row

    def change_log(self, team_id, schedule_id=None, limit=100):
        rows = [r for r in self.changes if r["team_id"] == team_id]
        if schedule_id:
            rows = [r for r in rows if r["schedule_id"] == schedule_id]
        return list(reversed(rows))[:limit]

    def list_requests(self, team_id, employee=None, status=None):
        rows = [r for r in self.requests if r["team_id"] == team_id]
        if status:
            rows = [r for r in rows if r["status"] == status]
        return rows


class _ScriptedLlm:
    def __init__(self, answers=None):
        self._answers = list(answers or [])
        self.calls = []

    def complete_json(self, system, user, schema=None, flow=""):
        self.calls.append({"system": system, "user": user, "schema": schema})
        if not self._answers:
            raise AssertionError("model called more times than scripted")
        answer = self._answers.pop(0)
        if isinstance(answer, Exception):
            raise answer
        return answer


def _generation(assignments, notes=None):
    return {"assignments": assignments, "notes": notes or [],
            "summary": "סידור השבוע"}


def _build_app(answers=None, launch=None, settings=None):
    repository = _FakeScheduleRepo()
    llm = _ScriptedLlm(answers)
    repository.model_calls = llm.calls
    guards = Guards(SECRET)
    app = FastAPI()
    app.include_router(
        schedules.build_router(
            ScheduleService(
                repository, llm, launch=launch, settings=settings,
                # The retry backoff is real seconds in production and nothing
                # here: these tests describe *that* a span is re-asked, which
                # is not the same claim as how long the pause is.
                sleep=lambda _seconds: None,
            ),
            guards,
        )
    )

    # The real handler's body builder, not a copy of it. A second shaping of
    # the same response is how a field the API actually returns goes missing
    # under test -- the tests would agree with themselves and disagree with
    # `app/main.py`.
    @app.exception_handler(AppError)
    async def handler(request, exc):
        return JSONResponse(
            status_code=exc.status_code, content=error_payload(exc)
        )

    return app, repository


def _client(app, role=ROLE_BOSS, team=TEAM):
    client = TestClient(app)
    client.cookies.set(COOKIE_NAME, issue(SECRET, team, role, 30))
    return client


def test_day_build_falls_back_when_the_decision_agent_is_unavailable():
    app, repo = _build_app([])
    client = _client(app)
    opened = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-23", "ends_on": "2026-08-29",
    }).json()

    proposal = client.post("/api/schedule/propose", json={
        "request": "תשבץ את שבת", "schedule_id": opened["id"],
    })
    assert proposal.status_code == 200
    body = proposal.json()
    assert body["operations"][0]["action"] == "generate_day"
    assert body["operations"][0]["previewed"] is True
    assert body["operations"][0]["assignments"]
    assert len(repo.model_calls) == 1

    applied = client.post("/api/schedule/apply", json={
        "schedule_id": opened["id"],
        "operations": body["operations"],
        "reason": "",
        "agent_reason": body["agent_reason"],
    })
    assert applied.status_code == 200
    assert applied.json()["assignments"]
    assert len(repo.model_calls) == 1


def test_agent_sees_candidates_and_its_choice_is_the_confirmed_schedule():
    app, repo = _build_app([{
        "candidate": 1,
        "reply": "הרצתי את השיבוץ ובחרתי בחלופה המאוזנת לאישור.",
        "agent_reason": "בחרתי ביוסי מתוך שתי החלופות החוקיות.",
    }])
    client = _client(app)
    opened = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-23", "ends_on": "2026-08-29",
    }).json()

    proposal = client.post("/api/schedule/propose", json={
        "request": "תשבץ את שבת", "schedule_id": opened["id"],
    })
    assert proposal.status_code == 200
    body = proposal.json()
    operation = body["operations"][0]
    assert operation["previewed"] is True
    assert [row["employee"] for row in operation["assignments"]] == ["יוסי"]
    assert body["agent_reason"] == "בחרתי ביוסי מתוך שתי החלופות החוקיות."
    assert len(repo.model_calls) == 1

    applied = client.post("/api/schedule/apply", json={
        "schedule_id": opened["id"],
        "operations": body["operations"],
        "reason": "",
        "agent_reason": body["agent_reason"],
    })
    assert applied.status_code == 200
    assert [row["employee"] for row in applied.json()["assignments"]] == [
        "יוסי"
    ]
    assert len(repo.model_calls) == 1


def test_a_generated_preview_cannot_be_rewritten_before_confirmation():
    app, _repo = _build_app([{
        "candidate": 0,
        "reply": "החלופה מוכנה לאישור.",
        "agent_reason": "החלופה חוקית ומאוזנת.",
    }])
    client = _client(app)
    opened = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-23", "ends_on": "2026-08-29",
    }).json()
    body = client.post("/api/schedule/propose", json={
        "request": "תשבץ את שבת", "schedule_id": opened["id"],
    }).json()
    body["operations"][0]["assignments"][0]["employee"] = "עובד שהומצא"

    applied = client.post("/api/schedule/apply", json={
        "schedule_id": opened["id"],
        "operations": body["operations"],
        "reason": "",
        "agent_reason": body["agent_reason"],
    })

    assert applied.status_code == 502
    assert "יש לבקש הצעה חדשה" in applied.json()["detail"]


def test_manual_assignment_cannot_cross_a_mandatory_round():
    app, repo = _build_app([])
    repo.profiles[TEAM] = {
        "workplace": {
            "name": "יחידה",
            "round_first_closure_date": "2026-08-29",
            "round_first_closure_group": "א",
        },
        "employees": [
            {"name": "סבב א", "exit_pattern": "round", "rotation_group": "א"},
            {"name": "סבב ב", "exit_pattern": "round", "rotation_group": "ב"},
        ],
        "shifts": PROFILE["shifts"],
        "rules": [],
    }
    client = _client(app)
    opened = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-29", "ends_on": "2026-08-29",
    }).json()

    blocked = client.post("/api/schedule/assign", json={
        "schedule_id": opened["id"], "employee": "סבב ב",
        "shift_name": MORNING, "slot_date": "2026-08-29",
    })
    assert blocked.status_code == 502
    assert "סבב או תלתון" in blocked.json()["detail"]

    allowed = client.post("/api/schedule/assign", json={
        "schedule_id": opened["id"], "employee": "סבב א",
        "shift_name": MORNING, "slot_date": "2026-08-29",
    })
    assert allowed.status_code == 200


class _DeferredLauncher:
    def __init__(self):
        self.jobs = []

    def __call__(self, target, *args):
        self.jobs.append((target, args))

    def run_next(self):
        target, args = self.jobs.pop(0)
        target(*args)


# -- generating ------------------------------------------------------------

def test_generating_stores_a_draft_with_reasons():
    app, repo = _build_app([])
    response = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "draft"
    assert "זמינות" in body["assignments"][0]["reason"]
    assert repo.model_calls == []


def test_generation_does_not_wait_for_a_model_to_read_soft_preferences():
    app, repo = _build_app([])
    repo.preference_rows = [
        {
            "team_id": TEAM, "status": "active", "kind": "employee",
            "subject": "דנה", "text": "דנה מעדיפה בקרים",
        },
        {
            "team_id": TEAM, "status": "suggested", "kind": "employee",
            "subject": "יוסי", "text": "אולי יוסי יעדיף ערב",
        },
    ]

    response = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    })

    assert response.status_code == 200
    assert response.json()["assignments"]
    assert repo.model_calls == []


def test_generating_enforces_the_managers_required_assignment():
    app, _ = _build_app([_generation([])])
    response = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17",
        "ends_on": "2026-08-18",
        "required_assignments": [{
            "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        }],
    })

    assert response.status_code == 200
    assert response.json()["assignments"][0]["employee"] == "דנה"
    assert "שיבוץ חובה" in response.json()["assignments"][0]["reason"]


def test_progressive_generation_supports_one_specific_date():
    app, _ = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "דנה מוסמכת לבוקר",
    }])])
    client = _client(app)

    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    assert started["generation"]["total_days"] == 1
    assert started["assignments"] == []

    finished = client.post(
        "/api/schedule/generate/%s/next" % started["id"]
    ).json()
    assert finished["generation"]["status"] == "complete"
    assert finished["generation"]["completed_days"] == 1
    assert finished["assignments"][0]["employee"] == "דנה"


def test_generation_run_returns_before_the_model_and_is_polled_with_get():
    launcher = _DeferredLauncher()
    app, repo = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "דנה מוסמכת לבוקר",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()

    queued = client.post(
        "/api/schedule/generate/%s/run" % started["id"]
    ).json()

    assert queued["generation"]["status"] == "running"
    assert repo.model_calls == []
    launcher.run_next()
    completed = client.get("/api/schedule/%s" % started["id"]).json()
    assert completed["generation"]["status"] == "complete"
    assert completed["assignments"][0]["employee"] == "דנה"


def test_generation_does_not_need_model_retries():
    """One bad answer costs a retry, not the rest of the period.

    The worker owns this, not the browser: nobody is watching a background
    build, so a job that parks itself on the first blip is a job that waits
    for a person who may not come back for an hour.
    """
    launcher = _DeferredLauncher()
    app, repo = _build_app([
        AgentError("תקלה זמנית"),
        _generation([{
            "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
            "reason": "הניסיון החוזר הצליח",
        }]),
    ], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()

    client.post("/api/schedule/generate/%s/run" % started["id"])
    launcher.run_next()

    generation = client.get(
        "/api/schedule/%s" % started["id"]
    ).json()["generation"]
    assert generation["status"] == "complete"
    assert generation["days"][0]["attempts"] == 1
    assert repo.model_calls == []


def test_failed_background_generation_requeues_as_running_for_polling():
    """Once the retries are spent the job parks, and `/run` resumes it."""
    launcher = _DeferredLauncher()
    app, repo = _build_app(
        # One more failure than the worker will absorb, so the job reaches
        # `failed` rather than recovering on its own.
        [AgentError("תקלה זמנית")] * 3 + [
            _generation([{
                "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
                "reason": "הניסיון החוזר הצליח",
            }]),
        ],
        launch=launcher,
    )
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    path = "/api/schedule/generate/%s/run" % started["id"]

    client.post(path)
    launcher.run_next()
    assert client.get("/api/schedule/%s" % started["id"]).json()[
        "generation"
    ]["status"] == "complete"
    assert repo.model_calls == []


def test_one_existing_day_can_be_rebuilt_with_board_instructions():
    launcher = _DeferredLauncher()
    app, repo = _build_app([_generation([{
        "employee": "יוסי", "shift": MORNING, "date": "2026-08-17",
        "reason": "יוסי בקבוצת הסגירה של שבת",
    }])], launch=launcher)
    client = _client(app)
    opened = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()

    prepared = client.post(
        "/api/schedule/generate/%s/day/start" % opened["id"],
        json={
            "date": "2026-08-17",
            "instructions": "זה יום הסגירה של קבוצה א; לשמור על הסבב",
        },
    ).json()
    client.post("/api/schedule/generate/%s/run" % opened["id"])

    assert prepared["generation"]["total_days"] == 1
    assert repo.model_calls == []
    launcher.run_next()
    completed = client.get("/api/schedule/%s" % opened["id"]).json()
    assert completed["assignments"]
    assert repo.schedules[opened["id"]]["generation"]["instructions"].startswith(
        "זה יום"
    )
    assert repo.model_calls == []


def test_progressive_long_range_is_one_persisted_request_per_day():
    app, _ = _build_app([
        _generation([{
            "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
            "reason": "כיסוי יום ראשון",
        }]),
        _generation([{
            "employee": "יוסי", "shift": MORNING, "date": "2026-08-18",
            "reason": "כיסוי יום שני",
        }]),
    ])
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
        "instructions": "לשמור על חלוקה הוגנת",
    }).json()

    first = client.post(
        "/api/schedule/generate/%s/next" % started["id"]
    ).json()
    assert first["generation"]["completed_days"] == 1
    assert first["generation"]["status"] == "running"

    second = client.post(
        "/api/schedule/generate/%s/next" % started["id"]
    ).json()
    assert second["generation"]["status"] == "complete"
    assert second["generation"]["completed_days"] == 2
    assert {row["date"] for row in second["assignments"]} == {
        "2026-08-17", "2026-08-18",
    }


def test_a_day_completes_even_when_the_model_would_fail():
    app, repo = _build_app([
        AgentError("תקלה זמנית"),
        _generation([{
            "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
            "reason": "ניסיון חוזר הצליח",
        }]),
    ])
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    path = "/api/schedule/generate/%s/next" % started["id"]

    completed = client.post(path)
    assert completed.status_code == 200
    assert completed.json()["generation"]["status"] == "complete"
    assert completed.json()["generation"]["days"][0]["attempts"] == 1
    assert repo.model_calls == []


def test_progress_answers_the_poll_without_rebuilding_the_grid():
    """The poll is asked once a second; it must not cost a full period.

    `GET /{id}` carries every slot, every assignment and a fresh audit over
    both. This route carries the counter, which is the entire question the
    browser is asking while a period is being built."""
    launcher = _DeferredLauncher()
    app, _ = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "דנה מוסמכת לבוקר",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    client.post("/api/schedule/generate/%s/run" % started["id"])

    response = client.get("/api/schedule/%s/progress" % started["id"])

    assert response.status_code == 200
    body = response.json()
    assert body["id"] == started["id"]
    assert body["generation"]["status"] == "running"
    # The counter and nothing else: no grid, no audit.
    assert "assignments" not in body and "warnings" not in body
    launcher.run_next()
    assert client.get(
        "/api/schedule/%s/progress" % started["id"]
    ).json()["generation"]["status"] == "complete"


def test_a_running_job_says_when_a_worker_last_touched_it():
    """What separates a slow model from a hung one.

    With no LLM timeout configured both look like "running" forever. A beat
    is the difference, and it is what lets the browser stop guessing."""
    launcher = _DeferredLauncher()
    app, repo = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "דנה מוסמכת לבוקר",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()

    assert started["generation"]["heartbeat"]
    client.post("/api/schedule/generate/%s/run" % started["id"])
    launcher.run_next()
    # Every checkpoint stamps one too, so the beat moves with the work even
    # between the worker's own timed beats.
    assert repo.schedules[started["id"]]["generation"]["heartbeat"]


def test_a_manager_can_stop_a_build_and_keep_the_finished_days():
    """Giving up is the manager's decision, and it costs them nothing.

    The days already paid for stay, the period is an ordinary draft
    immediately, and resuming picks up the first day that is not complete."""
    app, repo = _build_app([
        _generation([{
            "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
            "reason": "כיסוי יום ראשון",
        }]),
        _generation([{
            "employee": "יוסי", "shift": MORNING, "date": "2026-08-18",
            "reason": "כיסוי יום שני",
        }]),
    ])
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    client.post("/api/schedule/generate/%s/next" % started["id"])

    stopped = client.post(
        "/api/schedule/generate/%s/cancel" % started["id"]
    ).json()

    assert stopped["generation"]["status"] == "cancelled"
    assert stopped["generation"]["completed_days"] == 1
    assert [row["date"] for row in stopped["assignments"]] == ["2026-08-17"]
    assert stopped["status"] == "draft"


def test_a_stopped_job_takes_no_further_model_call():
    """Cooperative, but immediate at the next day boundary.

    The scripted model raises if it is called more times than scripted, so a
    worker that ignored the stop would fail this rather than pass it."""
    launcher = _DeferredLauncher()
    app, repo = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "כיסוי יום ראשון",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-19",
    }).json()
    client.post("/api/schedule/generate/%s/run" % started["id"])
    client.post("/api/schedule/generate/%s/cancel" % started["id"])

    launcher.run_next()

    assert repo.model_calls == []
    assert client.get("/api/schedule/%s/progress" % started["id"]).json()[
        "generation"
    ]["status"] == "cancelled"


def test_a_stopped_job_resumes_from_the_first_unfinished_day():
    launcher = _DeferredLauncher()
    app, _ = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "כיסוי יום ראשון",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    client.post("/api/schedule/generate/%s/cancel" % started["id"])

    resumed = client.post(
        "/api/schedule/generate/%s/run" % started["id"]
    ).json()

    assert resumed["generation"]["status"] == "running"
    assert resumed["generation"]["cancel_requested"] is False
    launcher.run_next()
    assert client.get("/api/schedule/%s/progress" % started["id"]).json()[
        "generation"
    ]["status"] == "complete"


def test_polling_the_same_job_twice_does_not_start_a_second_worker():
    """Two tabs, one job. A second worker would generate every day twice."""
    launcher = _DeferredLauncher()
    app, _ = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "כיסוי יום ראשון",
    }])], launch=launcher)
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    path = "/api/schedule/generate/%s/run" % started["id"]

    client.post(path)
    client.post(path)

    assert len(launcher.jobs) == 1


def test_a_hand_placed_shift_survives_the_day_being_generated():
    """The board stays writable while a period is built (D18), so a cell
    filled in mid-build must not be undone by the day that follows it.

    It goes to the model as a pin, and it is kept whether or not the model
    repeats it -- with the manager's own source, so the history still says a
    person put it there."""
    app, repo = _build_app([_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "בחירת הסוכן",
    }])])
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "יוסי", "schedule_id": started["id"],
    })

    built = client.post(
        "/api/schedule/generate/%s/next" % started["id"]
    ).json()

    placed = {row["employee"]: row["source"] for row in built["assignments"]}
    assert placed["יוסי"] == "manager"
    assert repo.model_calls == []


class _BlockingLlm:
    """A model that answers only when the test lets it.

    Stands in for the case the whole heartbeat/cancel machinery exists for:
    `llm_timeout_seconds` defaults to no limit, so a server that never
    answers holds its day open forever and the job stays `running` with
    nothing to read anywhere.
    """

    def __init__(self, answer):
        self._answer = answer
        self.entered = threading.Event()
        self.release = threading.Event()
        self.calls = []

    def complete_json(self, system, user, schema=None, flow=""):
        self.calls.append({"system": system, "user": user, "schema": schema})
        self.entered.set()
        self.release.wait(10)
        return self._answer


def test_a_model_that_never_answers_cannot_hold_generation_open():
    repository = _FakeScheduleRepo()
    llm = _BlockingLlm(_generation([{
        "employee": "דנה", "shift": MORNING, "date": "2026-08-17",
        "reason": "דנה מוסמכת לבוקר",
    }]))
    repository.model_calls = llm.calls
    guards = Guards(SECRET)
    app = FastAPI()
    app.include_router(
        schedules.build_router(ScheduleService(repository, llm), guards)
    )

    @app.exception_handler(AppError)
    async def handler(request, exc):
        return JSONResponse(
            status_code=exc.status_code, content=error_payload(exc)
        )

    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    client.post("/api/schedule/generate/%s/run" % started["id"])
    deadline = time.time() + 2
    while time.time() < deadline:
        state = client.get("/api/schedule/%s/progress" % started["id"]).json()
        if state["generation"]["status"] == "complete":
            break
        time.sleep(0.01)
    assert state["generation"]["status"] == "complete"
    assert not llm.entered.is_set()
    assert repository.model_calls == []


def _two_periods(client):
    """An older period and a newer one. The newer is what the server calls
    "current", so the older stands in for any week the manager paged to."""
    older = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-17",
    }).json()
    newer = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-24", "ends_on": "2026-08-24",
    }).json()
    return older, newer


def test_a_hand_placed_shift_lands_on_the_period_it_names():
    """The board is not always on the week that covers today.

    Every hand-write carries the period it happened on. Without one the
    server resolves "the period covering today", so a placement made on any
    other week went looking for its slot in the wrong period and came back
    404 -- while the placement check beside it, the one call that always sent
    the id, had just said the cell was fine."""
    app, _ = _build_app([])
    client = _client(app)
    older, newer = _two_periods(client)

    placed = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "schedule_id": older["id"],
    })

    assert placed.status_code == 200
    assert placed.json()["id"] == older["id"]
    assert [row["employee"] for row in placed.json()["assignments"]] == ["דנה"]
    # And the period the server would have guessed is untouched.
    assert client.get(
        "/api/schedule/%s" % newer["id"]
    ).json()["assignments"] == []


def test_a_drag_on_another_period_is_not_refused():
    """`move` had no `schedule_id` at all -- it resolved the current period
    unconditionally, so a drag on any other week could never find its target
    slot. The confirmation dialog now sends the period the drag happened
    on."""
    app, _ = _build_app([])
    client = _client(app)
    older, _ = _two_periods(client)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    placed = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "schedule_id": older["id"],
    }).json()
    row = placed["assignments"][0]

    moved = client.post("/api/schedule/move", json={
        "assignment_id": row["id"], "shift_name": MORNING,
        "slot_date": "2026-08-17", "reason": "החלפה מוסכמת",
        "schedule_id": older["id"],
    })

    assert moved.status_code == 200
    assert moved.json()["id"] == older["id"]


def test_removing_a_shift_names_its_period_too():
    app, _ = _build_app([])
    client = _client(app)
    older, _ = _two_periods(client)
    placed = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "schedule_id": older["id"],
    }).json()

    removed = client.post("/api/schedule/unassign", json={
        "assignment_id": placed["assignments"][0]["id"],
        "reason": "בוטל",
        "schedule_id": older["id"],
    })

    assert removed.status_code == 200
    assert removed.json()["assignments"] == []


def test_a_write_naming_no_period_still_means_the_current_one():
    """The fallback stays, for any client that predates the field."""
    app, _ = _build_app([])
    client = _client(app)
    _, newer = _two_periods(client)

    placed = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-24", "employee": "דנה",
    })

    assert placed.status_code == 200
    assert placed.json()["id"] == newer["id"]


def test_generating_without_a_finished_interview_is_refused():
    """No profile means no shift vocabulary, and guessing one is exactly
    the hardcoding D9 forbids."""
    app, repo = _build_app([])
    repo.profiles[TEAM] = None
    response = _client(app).post("/api/schedule/generate", json={})
    assert response.status_code == 502
    assert "ראיון" in response.json()["detail"]


def _days(first, last):
    return [
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-%02d" % day,
         "reason": "כיסוי"}
        for day in range(first, last + 1)
    ]


def test_a_generated_schedule_carries_its_warnings_and_still_returns_200():
    """D3: warnings are advisory. A schedule that breaks a rule still renders.

    Eight days is past `_CHUNK_DAYS`, so this also covers the case the warning
    needs most: a run of consecutive shifts that crosses a chunk boundary and
    is only visible once the chunks are merged. The audit sees the period, not
    the calls it took to build it.
    """
    app, _ = _build_app([
        _generation(_days(17, 23)), _generation(_days(24, 24)),
    ])
    response = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-24",
    })
    assert response.status_code == 200
    codes = {warning["code"] for warning in response.json()["warnings"]}
    # The deterministic engine avoids a provable consecutive-days breach
    # instead of accepting it from a model and merely reporting it later.
    assert "consecutive" not in codes
    assert len(response.json()["assignments"]) == 8


# -- the management overview ----------------------------------------------

def test_the_overview_returns_roster_vocabulary_and_period():
    app, _ = _build_app([_generation([
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17",
         "reason": "מוסמכת"},
    ])])
    client = _client(app)
    client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    body = client.get("/api/schedule/overview").json()
    assert [row["name"] for row in body["employees"]] == ["דנה", "יוסי"]
    assert [row["name"] for row in body["shifts"]] == [MORNING]
    assert body["schedule"]["assignments"]


def test_the_overview_carries_the_periods_numbers():
    """The charts' figures ride along on the overview the screen already
    fetches, so the panel and the calendar beside it can never be a refresh
    apart. Pinned through the HTTP contract because the Pydantic models are
    what the browser actually receives -- a field added to `shift_stats` but
    missing from `ShiftStats` would be silently dropped here.
    """
    app, _ = _build_app([_generation([
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17",
         "reason": "מוסמכת"},
    ])])
    client = _client(app)
    client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })

    stats = client.get("/api/schedule/overview").json()["stats"]

    assert stats["total_shifts"] == 2
    assert stats["people_working"] == 2
    # Everyone on the roster is present, including the person with no
    # shifts -- that zero is the whole point of the per-person chart.
    assert {row["employee"] for row in stats["by_employee"]} == {"דנה", "יוסי"}
    assert stats["coverage"]["assigned"] == 2


def test_the_overview_carries_zeroed_stats_before_anything_is_built():
    """The state the management screen opens in.

    The panel renders whatever arrives, so "no schedule yet" has to be a
    well-formed shape rather than a null the client special-cases.
    """
    app, _ = _build_app([])

    stats = _client(app).get("/api/schedule/overview").json()["stats"]

    assert stats["total_shifts"] == 0
    assert stats["coverage"]["required"] == 0


def test_a_member_sees_only_published_periods():
    """A draft is the manager's working state; publishing makes it the team's."""
    app, _ = _build_app([_generation([])])
    boss = _client(app)
    created = boss.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    member = _client(app, role=ROLE_MEMBER)
    assert member.get("/api/schedule/overview").json()["schedule"] is None

    boss.post("/api/schedule/%s/publish" % created["id"])
    assert member.get("/api/schedule/overview").json()["schedule"] is not None


# -- D18: the manual path -------------------------------------------------

def test_opening_a_blank_period_calls_no_model():
    """The authoring half of D6. `_ScriptedLlm` raises if it is called at
    all, so an empty answer list is the assertion: building a grid is
    arithmetic over the declared vocabulary, not a generation."""
    app, _ = _build_app([])
    response = _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    assert response.status_code == 200
    body = response.json()
    assert body["status"] == "draft"
    # The grid exists; nobody is on it yet.
    assert body["slots"], "a blank period still has its slots"
    assert body["assignments"] == []


def test_a_blank_period_without_a_finished_interview_is_refused():
    """Skipping the agent does not mean skipping the interview: without the
    shift vocabulary there is no grid to build (D9)."""
    app, repo = _build_app([])
    repo.profiles[TEAM] = None
    response = _client(app).post("/api/schedule/blank", json={})
    assert response.status_code == 502
    assert "ראיון" in response.json()["detail"]


def test_a_manual_assignment_is_marked_as_the_managers_and_calls_no_model():
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    response = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "reason": "דנה ביקשה את הבוקר הזה",
    })
    assert response.status_code == 200
    rows = response.json()["assignments"]
    assert len(rows) == 1
    assert rows[0]["employee"] == "דנה"
    # D18: provenance is recorded, so the agent and the manager are
    # distinguishable later rather than only by how their prose reads.
    assert rows[0]["source"] == "manager"
    assert rows[0]["reason"] == "דנה ביקשה את הבוקר הזה"


def test_a_manual_assignment_without_a_reason_still_carries_one():
    """D8 is answered by a different voice, not relaxed. The row says a
    person placed it rather than inventing a judgment the agent never made."""
    app, _ = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    response = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17", "employee": "דנה",
    })
    assert response.status_code == 200
    assert response.json()["assignments"][0]["reason"].strip()


def test_a_manual_assignment_is_recorded_in_the_change_log():
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "reason": "כיסוי בוקר",
    })
    actions = [row["action"] for row in repo.changes]
    assert "opened" in actions
    assert "assigned" in actions
    placed = [row for row in repo.changes if row["action"] == "assigned"][0]
    assert placed["employee"] == "דנה"
    assert placed["reason"] == "כיסוי בוקר"


def test_assigning_the_same_person_twice_changes_nothing():
    """The table is UNIQUE (slot_id, employee) and the insert does nothing on
    conflict, so a double click is a success that added no duplicate."""
    app, _ = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    body = {"shift_name": MORNING, "slot_date": "2026-08-17",
            "employee": "דנה"}
    first = client.post("/api/schedule/assign", json=body).json()
    second = client.post("/api/schedule/assign", json=body).json()
    assert len(second["assignments"]) == 1
    assert first["assigned"] == second["assigned"]


def test_assigning_onto_a_shift_that_does_not_run_is_a_404():
    app, _ = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    response = client.post("/api/schedule/assign", json={
        "shift_name": EVENING, "slot_date": "2026-08-17", "employee": "דנה",
    })
    assert response.status_code == 404


def test_a_manual_schedule_still_carries_its_warnings():
    """D3 holds on the manual path too: the audit reports, and a hand-built
    week that breaks a rule still returns 200 and still renders."""
    app, _ = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-24",
    })
    for day in range(17, 25):
        client.post("/api/schedule/assign", json={
            "shift_name": MORNING, "slot_date": "2026-08-%02d" % day,
            "employee": "דנה",
        })
    response = client.get("/api/schedule/overview")
    assert response.status_code == 200
    schedule = response.json()["schedule"]
    codes = {warning["code"] for warning in schedule["warnings"]}
    assert "consecutive" in codes
    # Reported, never withheld.
    assert len(schedule["assignments"]) == 8


def test_unassigning_removes_the_row_and_logs_it():
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    placed = client.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17", "employee": "דנה",
    }).json()
    assignment_id = placed["assignments"][0]["id"]

    response = client.post("/api/schedule/unassign", json={
        "assignment_id": assignment_id, "reason": "דנה התחלפה",
    })
    assert response.status_code == 200
    assert response.json()["assignments"] == []
    removed = [row for row in repo.changes if row["action"] == "removed"][0]
    assert removed["employee"] == "דנה"
    assert removed["reason"] == "דנה התחלפה"


def test_clearing_a_day_empties_it_and_leaves_the_rest_of_the_week():
    """The counterpart to building a week: taking one day back.

    Logged per row rather than as one "the day was cleared" entry — the log
    is the only history there is (D4), and a manager asking later which
    shifts those were has to be able to find out.
    """
    app, repo = _build_app([])
    client = _client(app)
    created = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    for day in ("2026-08-17", "2026-08-18"):
        client.post("/api/schedule/assign", json={
            "shift_name": MORNING, "slot_date": day, "employee": "דנה",
            "schedule_id": created["id"],
        })

    response = client.post("/api/schedule/%s/clear" % created["id"], json={
        "slot_date": "2026-08-17", "reason": "בונים את היום מחדש",
    })

    assert response.status_code == 200
    body = response.json()
    assert body["cleared"] == 1
    assert [row["date"] for row in body["assignments"]] == ["2026-08-18"]
    # The grid stays: the week's shape comes from the vocabulary, not from
    # what happened to be assigned into it.
    assert len(body["slots"]) == 2
    removed = [row for row in repo.changes if row["action"] == "removed"]
    assert [row["slot_date"] for row in removed] == ["2026-08-17"]
    assert removed[0]["reason"] == "בונים את היום מחדש"


def test_clearing_without_a_date_empties_the_whole_period():
    app, _ = _build_app([])
    client = _client(app)
    created = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    for day in ("2026-08-17", "2026-08-18"):
        client.post("/api/schedule/assign", json={
            "shift_name": MORNING, "slot_date": day, "employee": "דנה",
            "schedule_id": created["id"],
        })

    body = client.post(
        "/api/schedule/%s/clear" % created["id"], json={},
    ).json()

    assert body["cleared"] == 2
    assert body["assignments"] == []
    assert len(body["slots"]) == 2


def test_clearing_an_empty_day_reports_that_nothing_went():
    """Not a failure: the manager asked for an empty day and it is empty."""
    app, _ = _build_app([])
    client = _client(app)
    created = client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    body = client.post("/api/schedule/%s/clear" % created["id"], json={
        "slot_date": "2026-08-17",
    }).json()

    assert body["cleared"] == 0


def test_one_workspace_cannot_clear_anothers_period():
    app, _ = _build_app([])
    created = _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    intruder = _client(app, team=OTHER_TEAM)

    assert intruder.post(
        "/api/schedule/%s/clear" % created["id"], json={},
    ).status_code == 404


def test_unassigning_something_that_is_not_there_is_a_404():
    app, _ = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    response = client.post("/api/schedule/unassign", json={
        "assignment_id": "nope",
    })
    assert response.status_code == 404


def test_a_generated_assignment_is_marked_as_the_agents():
    """The other side of D18: nothing about the manual path changes what a
    generated row says about itself."""
    app, _ = _build_app([_generation([
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17",
         "reason": "דנה מוסמכת לבוקר"},
    ])])
    response = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    assert response.json()["assignments"][0]["source"] == "agent"


def test_one_workspace_cannot_assign_into_anothers_schedule():
    app, _ = _build_app([])
    created = _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    intruder = _client(app, team=OTHER_TEAM)
    response = intruder.post("/api/schedule/assign", json={
        "shift_name": MORNING, "slot_date": "2026-08-17",
        "employee": "דנה", "schedule_id": created["id"],
    })
    assert response.status_code == 404


# -- D5: employees are read-only ------------------------------------------

@pytest.mark.parametrize("method,path,body", [
    ("post", "/api/schedule/generate", {}),
    ("post", "/api/schedule/blank", {}),
    ("post", "/api/schedule/assign",
     {"shift_name": MORNING, "slot_date": "2026-08-17", "employee": "דנה"}),
    ("post", "/api/schedule/unassign", {"assignment_id": "a"}),
    ("post", "/api/schedule/propose", {"request": "דנה חולה"}),
    ("post", "/api/schedule/apply",
     {"schedule_id": "s", "operations": [], "reason": "מחלה"}),
    ("post", "/api/schedule/move",
     {"assignment_id": "a", "shift_name": MORNING,
      "slot_date": "2026-08-17", "reason": "מחלה"}),
    ("post", "/api/schedule/check",
     {"employee": "דנה", "shift_name": MORNING, "slot_date": "2026-08-17"}),
    ("post", "/api/schedule/constraints",
     {"employee": "דנה", "constraint_date": "2026-08-17"}),
    ("delete", "/api/schedule/constraints/x", None),
    ("post", "/api/schedule/some-id/clear", {}),
    ("delete", "/api/schedule/some-id", None),
])
def test_a_member_cannot_reach_any_mutation(method, path, body):
    """D5, enforced by `guards.boss()` rather than by convention.

    Parameterised over every mutating route so a route added later without a
    boss guard fails here rather than shipping.
    """
    app, _ = _build_app([])
    member = _client(app, role=ROLE_MEMBER)
    response = getattr(member, method)(
        path, **({"json": body} if body is not None else {})
    )
    assert response.status_code == 401


# -- D10: tenant isolation -------------------------------------------------

def test_another_workspace_cannot_read_a_schedule_by_id():
    app, _ = _build_app([_generation([])])
    created = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    intruder = _client(app, team=OTHER_TEAM)
    response = intruder.get("/api/schedule/%s" % created["id"])
    # 404, not 403: a distinct "wrong team" would confirm the id is real.
    assert response.status_code == 404


def test_another_workspace_cannot_publish_a_schedule_by_id():
    app, _ = _build_app([_generation([])])
    created = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    intruder = _client(app, team=OTHER_TEAM)
    assert intruder.post(
        "/api/schedule/%s/publish" % created["id"]
    ).status_code == 404


def test_constraints_are_scoped_to_the_workspace():
    app, _ = _build_app([])
    _client(app).post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-20",
        "reason": "מחלה",
    })
    intruder = _client(app, team=OTHER_TEAM)
    assert intruder.get("/api/schedule/constraints/list").json() == []


# -- the change loop -------------------------------------------------------

def test_a_change_without_a_reason_is_answered_by_asking():
    app, _ = _build_app([{
        "reply": "למה דנה לא מגיעה?", "needs_reason": True,
        "agent_reason": "", "operations": [], "constraints": [],
    }])
    client = _client(app)
    client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    proposal = client.post("/api/schedule/propose", json={
        "request": "תוריד את דנה",
    }).json()
    assert proposal["needs_reason"] is True
    assert proposal["operations"] == []


def test_proposing_persists_nothing():
    """The two-step contract: a proposal is a proposal until confirmed."""
    app, repo = _build_app([
        {"reply": "אציע להוריד את דנה", "needs_reason": False,
         "agent_reason": "דנה חולה, יוסי פנוי",
         "operations": [{"action": "remove", "employee": "דנה",
                         "shift": MORNING, "date": "2026-08-17",
                         "reason": "חולה"}],
         "constraints": []},
    ])
    client = _client(app)
    created = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    before = len(repo.assignment_rows[created["id"]])

    client.post("/api/schedule/propose", json={
        "request": "דנה חולה", "reason": "מחלה",
    })
    assert len(repo.assignment_rows[created["id"]]) == before


def test_applying_a_confirmed_change_records_both_reasons():
    """D8: the manager's reason and the agent's both land in the log."""
    app, repo = _build_app([
        _generation([{"employee": "דנה", "shift": MORNING,
                      "date": "2026-08-17", "reason": "מוסמכת"}]),
    ])
    client = _client(app)
    created = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    response = client.post("/api/schedule/apply", json={
        "schedule_id": created["id"],
        "operations": [{"action": "remove", "employee": "דנה",
                        "shift": MORNING, "date": "2026-08-17",
                        "reason": "דנה חולה"}],
        "reason": "מחלה",
        "agent_reason": "אין תחליף זמין",
    })
    assert response.status_code == 200
    entry = next(
        row for row in repo.changes if row["action"] == "removed"
    )
    assert entry["reason"] == "מחלה"
    assert entry["agent_reason"] == "דנה חולה"


def test_applying_without_a_reason_is_rejected_by_validation():
    """By apply time the manager has been asked; a blank reason now would be
    a hole in the only history the system keeps."""
    app, _ = _build_app([])
    response = _client(app).post("/api/schedule/apply", json={
        "schedule_id": "s", "operations": [], "reason": "",
    })
    assert response.status_code == 422


def test_agent_can_propose_and_apply_adding_an_employee_without_a_schedule():
    """Profile maintenance is conversational too, but still confirmed."""
    app, repo = _build_app([{
        "reply": "אוסיף את מאיה לצוות",
        "needs_reason": False,
        "agent_reason": "פרטי העובדת נמסרו במפורש",
        "operations": [],
        "constraints": [],
        "profile_operations": [{
            "action": "add_employee",
            "target": "",
            "item": {
                "name": "מאיה", "role": "אחראית משמרת",
                "eligible_shifts": [MORNING], "start_time": "",
                "end_time": "", "headcount": 1, "is_on_call": False,
            },
        }],
    }])
    client = _client(app)

    proposal = client.post("/api/schedule/propose", json={
        "request": "תוסיף את מאיה לצוות כאחראית משמרת",
    }).json()
    assert proposal["schedule_id"] == ""
    assert proposal["profile_operations"][0]["item"]["name"] == "מאיה"
    assert [row["name"] for row in repo.profiles[TEAM]["employees"]] == [
        "דנה", "יוסי",
    ]

    response = client.post("/api/schedule/apply", json={
        "profile_operations": proposal["profile_operations"],
    })
    assert response.status_code == 200
    assert [row["name"] for row in repo.profiles[TEAM]["employees"]] == [
        "דנה", "יוסי", "מאיה",
    ]


# -- the drag, which is a proposal ----------------------------------------

def test_a_confirmed_drag_moves_the_assignment_and_logs_the_reason():
    """Dragging is a way of proposing a change, not a way around D8.

    The gesture opens a confirmation; this is what that dialog sends, and it
    carries the manager's reason exactly as a spoken change does.
    """
    app, repo = _build_app([
        _generation([{"employee": "דנה", "shift": MORNING,
                      "date": "2026-08-17", "reason": "מוסמכת"}]),
    ])
    client = _client(app)
    created = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    assignment = created["assignments"][0]

    response = client.post("/api/schedule/move", json={
        "assignment_id": assignment["id"],
        "shift_name": MORNING,
        "slot_date": "2026-08-18",
        "reason": "דנה ביקשה להחליף יום",
    })
    assert response.status_code == 200
    moved = response.json()["assignments"][0]
    assert moved["date"] == "2026-08-18"
    entry = next(row for row in repo.changes if row["action"] == "moved")
    assert entry["reason"] == "דנה ביקשה להחליף יום"


def test_a_drag_without_a_reason_is_rejected():
    """The confirmation dialog is what collects it; an empty one never
    reaches the server."""
    app, _ = _build_app([])
    response = _client(app).post("/api/schedule/move", json={
        "assignment_id": "a", "shift_name": MORNING,
        "slot_date": "2026-08-18", "reason": "",
    })
    assert response.status_code == 422


# -- constraints -----------------------------------------------------------

def test_a_constraint_is_recorded_with_its_source():
    """`source` records where the information came from, not who typed it:
    employees have no account and never write here (D5/D10)."""
    app, _ = _build_app([])
    client = _client(app)
    response = client.post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-20",
        "reason": "תואר", "source": "employee_reported",
    })
    assert response.status_code == 200
    assert response.json()["source"] == "employee_reported"


def test_a_timed_soft_constraint_is_recorded():
    app, _ = _build_app([])
    response = _client(app).post("/api/schedule/constraints", json={
        "employee": "עודד", "constraint_date": "2026-08-20",
        "available": True, "start_time": "16:00", "is_hard": False,
        "reason": "לימודים",
    })

    assert response.status_code == 200
    assert response.json()["start_time"] == "16:00"
    assert response.json()["is_hard"] is False


def test_an_invalid_constraint_time_is_rejected():
    app, _ = _build_app([])
    response = _client(app).post("/api/schedule/constraints", json={
        "employee": "עודד", "constraint_date": "2026-08-20",
        "available": True, "start_time": "29:00",
    })
    assert response.status_code == 422


def test_a_recorded_constraint_shows_up_as_a_warning_when_contradicted():
    """The audit reads a constraint with no shift as covering the whole day."""
    app, _ = _build_app([_generation([
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17",
         "reason": "כיסוי"},
    ])])
    client = _client(app)
    client.post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-17",
        "reason": "מילואים",
    })
    body = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    assert not any(
        row["employee"] == "דנה" and row["date"] == "2026-08-17"
        for row in body["assignments"]
    )


def test_restating_a_constraint_replaces_rather_than_duplicates():
    """Two rows saying the same thing would double every warning it causes."""
    app, repo = _build_app([])
    client = _client(app)
    for reason in ("מחלה", "מילואים"):
        client.post("/api/schedule/constraints", json={
            "employee": "דנה", "constraint_date": "2026-08-20",
            "reason": reason,
        })
    rows = client.get("/api/schedule/constraints/list").json()
    assert len(rows) == 1
    assert rows[0]["reason"] == "מילואים"


def test_the_team_can_read_constraints_but_not_write_them():
    app, _ = _build_app([])
    _client(app).post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-20", "reason": "תואר",
    })
    member = _client(app, role=ROLE_MEMBER)
    assert len(member.get("/api/schedule/constraints/list").json()) == 1
    assert member.post("/api/schedule/constraints", json={
        "employee": "יוסי", "constraint_date": "2026-08-21",
    }).status_code == 401


def test_the_change_log_is_readable_as_history():
    app, _ = _build_app([_generation([])])
    client = _client(app)
    client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    history = client.get("/api/schedule/history/list").json()
    assert any(row["action"] == "generated" for row in history)


# -- the agent speaking first ----------------------------------------------

def test_the_agent_briefs_the_manager_unprompted():
    """The one route the manager did not initiate (D15)."""
    app, _ = _build_app([
        {"headline": "יש חור בשלישי", "quiet": False, "items": [
            {"text": "משמרת בוקר של שלישי ריקה", "kind": "gap",
             "suggestion": "מי יכול לכסות את בוקר שלישי?"},
        ]},
    ])
    client = _client(app)
    client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })

    body = client.post("/api/schedule/brief", json={
        "trigger": "opened", "last_said": [],
    }).json()

    assert body["quiet"] is False
    assert body["items"][0]["suggestion"]


def test_a_briefing_never_breaks_the_screen():
    """No model answer is scripted, so the call fails inside the service.

    It must come back quiet and `200`. This sits beside a calendar that has
    to render whatever the model is doing — an error here would take the
    management area down with it.
    """
    app, _ = _build_app([])

    response = _client(app).post("/api/schedule/brief", json={
        "trigger": "opened",
    })

    assert response.status_code == 200
    assert response.json() == {"headline": "", "items": [], "quiet": True}


def test_a_member_cannot_ask_for_a_briefing():
    """A briefing reads drafts and other people's stated reasons."""
    app, _ = _build_app([])
    member = _client(app, role=ROLE_MEMBER)

    assert member.post(
        "/api/schedule/brief", json={"trigger": "opened"}
    ).status_code == 401


def test_briefing_before_the_interview_is_quiet_without_calling_the_model():
    """Nothing to observe: the agent knows neither the shifts nor the people,
    and a briefing built on that would be invented rather than noticed."""
    app, repo = _build_app([])
    repo.profiles = {}

    body = _client(app).post(
        "/api/schedule/brief", json={"trigger": "opened"}
    ).json()

    assert body["quiet"] is True


# -- leaving the app as a file (D17) ---------------------------------------

def test_a_period_downloads_as_a_workbook():
    app, _ = _build_app([_generation([
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17",
         "reason": "דנה מוסמכת לבוקר"},
    ])])
    client = _client(app)
    schedule = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    response = client.get("/api/schedule/export/%s" % schedule["id"])

    assert response.status_code == 200
    assert "spreadsheetml" in response.headers["content-type"]
    assert "attachment" in response.headers["content-disposition"]
    # A real xlsx is a zip; anything else means an error page came back 200.
    assert response.content[:2] == b"PK"


def test_a_member_cannot_download_the_workbook():
    """The share link is how the team reads the roster. A file is a copy that
    leaves the app entirely, and handing that out is the manager's call."""
    app, _ = _build_app([_generation([])])
    client = _client(app)
    schedule = client.post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    member = _client(app, role=ROLE_MEMBER)
    assert member.get(
        "/api/schedule/export/%s" % schedule["id"]
    ).status_code == 401


def test_one_workspace_cannot_download_anothers_period():
    app, _ = _build_app([_generation([])])
    schedule = _client(app).post("/api/schedule/generate", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    other = _client(app, team=OTHER_TEAM)
    assert other.get(
        "/api/schedule/export/%s" % schedule["id"]
    ).status_code == 404


# -- import (D7) -----------------------------------------------------------

def _upload(book, name="schedule.xlsx"):
    import io

    stream = io.BytesIO()
    book.save(stream)
    return (
        "files",
        (name, stream.getvalue(),
         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
    )


def _sheet(grid):
    """An ad-hoc `.xlsx` from a list of rows, for one-off layouts."""
    from openpyxl import Workbook

    book = Workbook()
    for row in grid:
        book.active.append(row)
    return book


def _preview_grid(client, grid):
    return client.post(
        "/api/schedule/import/preview?learn_rules=false",
        files=[_upload(_sheet(grid), "sheet.xlsx")],
    ).json()


def _preview(client, books, learn_rules=False):
    return client.post(
        "/api/schedule/import/preview?learn_rules=%s"
        % ("true" if learn_rules else "false"),
        files=[_upload(book, "file-%d.xlsx" % index)
               for index, book in enumerate(books)],
    )


def test_previewing_an_import_reads_the_file():
    from tests.fixtures.build import sample_a

    app, repo = _build_app([])
    response = _preview(_client(app), [sample_a()])
    assert response.status_code == 200
    body = response.json()
    assert body["periods"][0]["layout"] == "shift_major"
    # Both rows are read. `PROFILE` declares only `בוקר`, and the `צהריים`
    # row is still taken from the sheet's own wording rather than dropped --
    # the file records something the workplace really ran.
    assert len(body["periods"][0]["assignments"]) == 10


def test_a_shift_the_workplace_never_declared_is_still_read():
    """A file may name a shift the interview never heard of (D7).

    Sample A has a `צהריים` row this workplace never declared. Refusing it
    would silently lose real history and would make the product reject the
    very files it exists to absorb. Nothing is invented — the name is the
    manager's own.
    """
    from tests.fixtures.build import sample_a

    app, repo = _build_app([])
    body = _preview(_client(app), [sample_a()]).json()
    assert body["periods"][0]["shifts"] == [MORNING, EVENING]


def test_a_declared_shift_is_still_matched_to_its_declared_name():
    """D9 where it applies: the vocabulary's spelling wins."""
    app, repo = _build_app([])
    body = _preview_grid(_client(app), [
        ["משמרות", "1/6/25"],
        ["משמרת בוקר", "דנה"],
    ])
    assert body["periods"][0]["shifts"] == [MORNING]


def test_a_sheet_of_only_dates_and_people_is_read():
    """No shift column at all — the plainest file a manager keeps."""
    app, repo = _build_app([])
    body = _preview_grid(_client(app), [
        ["", "1/6/25", "2/6/25"],
        ["", "דנה", "יוסי"],
    ])
    period = body["periods"][0]
    assert period["layout"] == "date_only"
    assert period["shifts"] == []
    assert any("אין שמות משמרות" in note for note in period["warnings"])


def test_previewing_an_import_stores_nothing():
    """D7's whole point: inference is not a write.

    The confirmation is only real if the preview has not already committed.
    """
    from tests.fixtures.build import sample_a

    app, repo = _build_app([])
    _preview(_client(app), [sample_a()])
    assert repo.schedules == {}
    assert repo.availability_rows == []
    assert repo.changes == []


def test_confirming_an_import_stores_the_approved_rows():
    app, repo = _build_app([])
    response = _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
            {"employee": "יוסי", "shift": MORNING, "date": "2026-08-18"},
        ],
    })
    assert response.status_code == 200
    assert len(response.json()["assignments"]) == 2
    assert len(repo.schedules) == 1


def test_an_imported_row_says_it_was_imported():
    """`assignments.source` distinguishes a recorded past from a decision."""
    app, repo = _build_app([])
    _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
        ],
    })
    rows = list(repo.assignment_rows.values())[0]
    assert rows[0]["source"] == "imported"


def test_an_imported_row_carries_a_reason_without_inventing_one():
    """D8 is answered by a different voice, not relaxed."""
    app, repo = _build_app([])
    _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
        ],
    })
    rows = list(repo.assignment_rows.values())[0]
    assert rows[0]["reason"]


def test_a_stated_constraint_is_stored_as_the_managers_own():
    """Nobody submitted it, so `employee_reported` would be a lie (D13)."""
    app, repo = _build_app([])
    _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
        ],
        "unavailability": [
            {"employee": "יוסי", "date": "2026-08-18", "shift": EVENING,
             "reason": "לא זמין"},
        ],
    })
    assert repo.availability_rows[0]["source"] == "manager"
    assert repo.availability_rows[0]["available"] is False


def test_the_grid_comes_from_the_file_not_from_todays_vocabulary():
    """A past schedule ran the shifts it ran.

    `EVENING` is not in `PROFILE`'s shift list, so rebuilding the grid with
    `build_slots` would silently drop it and reshape history.
    """
    app, repo = _build_app([])
    _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "יוסי", "shift": EVENING, "date": "2026-08-17"},
        ],
    })
    slots = list(repo.slots.values())[0]
    assert [slot["shift_name"] for slot in slots] == [EVENING]


def test_an_unreadable_file_does_not_sink_the_batch():
    """A folder of a year's sheets will contain one stray document."""
    from tests.fixtures.build import sample_a

    app, repo = _build_app([])
    response = _client(app).post(
        "/api/schedule/import/preview?learn_rules=false",
        files=[
            _upload(sample_a(), "good.xlsx"),
            ("files", ("junk.xlsx", b"not a spreadsheet",
                       "application/octet-stream")),
        ],
    )
    assert response.status_code == 200
    body = response.json()
    assert len(body["periods"]) == 1
    assert body["failures"][0]["filename"] == "junk.xlsx"


def test_a_workbook_reads_schedule_tabs_even_when_summary_is_active():
    from openpyxl import Workbook

    book = Workbook()
    book.active.title = "סיכום"
    book.active.append(["מדד", "ערך"])
    schedule = book.create_sheet("שבוע 1")
    schedule.append(["משמרות", "1/6/25"])
    schedule.append([MORNING, "דנה"])

    app, repo = _build_app([])
    body = _preview(_client(app), [book]).json()

    assert len(body["periods"]) == 1
    assert body["periods"][0]["filename"].endswith("שבוע 1")
    assert body["periods"][0]["assignments"][0]["employee"] == "דנה"
    assert body["failures"][0]["filename"].endswith("סיכום")


def test_no_readable_file_at_all_is_an_error():
    app, repo = _build_app([])
    response = _client(app).post(
        "/api/schedule/import/preview?learn_rules=false",
        files=[("files", ("junk.xlsx", b"nope", "application/octet-stream"))],
    )
    assert response.status_code >= 400


def test_learned_rules_are_never_returned_pre_approved():
    """A candidate becomes a rule only when the manager says so (D7)."""
    from tests.fixtures.build import sample_a

    app, repo = _build_app([{
        "rules": [{"text": "אלמוג עובד בעיקר בוקר", "kind": "soft",
                   "evidence": "2 מתוך 3", "confidence": "medium"}],
        "notes": [],
    }])
    response = _preview(_client(app), [sample_a()], learn_rules=True)
    assert response.status_code == 200
    rules = response.json()["candidate_rules"]
    assert rules and rules[0]["approved"] is False


def test_patterns_are_counted_across_all_the_uploaded_files():
    """A pattern is by definition what one period cannot show."""
    from tests.fixtures.build import sample_a, sample_b

    app, repo = _build_app([])
    repo.profiles[TEAM] = dict(PROFILE, shifts=[
        {"name": MORNING}, {"name": EVENING},
    ])
    single = _preview(_client(app), [sample_a()]).json()
    both = _preview(_client(app), [sample_a(), sample_b()]).json()
    # The tally spans the batch: Sample B's rows are counted alongside
    # Sample A's, which is the only way a cross-period pattern is visible.
    assert (both["observations"]["totals"]["assignments"]
            > single["observations"]["totals"]["assignments"])
    assert len(both["periods"]) == 2


def test_a_shiftless_row_cannot_be_confirmed():
    """The read shape allows an empty shift; the stored shape does not.

    A sheet of dates and people carries no shift, and that stays visible
    through the preview. By the time rows are being stored the manager has
    been asked, and a shiftless assignment has no slot on the grid to sit in.
    """
    app, repo = _build_app([])
    response = _client(app).post("/api/schedule/import/confirm", json={
        "assignments": [
            {"employee": "דנה", "shift": "", "date": "2026-08-17"},
        ],
    })
    assert response.status_code == 422
    assert repo.schedules == {}


def test_a_member_cannot_preview_an_import():
    from tests.fixtures.build import sample_a

    app, repo = _build_app([])
    response = _preview(_client(app, role=ROLE_MEMBER), [sample_a()])
    assert response.status_code in (401, 403)


def test_a_member_cannot_confirm_an_import():
    app, repo = _build_app([])
    response = _client(app, role=ROLE_MEMBER).post(
        "/api/schedule/import/confirm", json={
            "assignments": [
                {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
            ],
        })
    assert response.status_code in (401, 403)
    assert repo.schedules == {}


# -- the board: checking a placement before making it ----------------------

def test_checking_a_placement_calls_no_model():
    """The board's validation runs with the agent unavailable.

    `_ScriptedLlm` raises the moment it is called with nothing scripted, so
    an empty script is the assertion: if `/check` ever grew a model call,
    this test fails rather than quietly getting slower. That property is the
    whole reason `bl/placement.py` exists separately from `changes.py`.
    """
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    response = client.post("/api/schedule/check", json={
        "employee": "דנה", "shift_name": MORNING, "slot_date": "2026-08-17",
    })
    assert response.status_code == 200
    assert response.json()["ok"] is True


def test_a_check_that_warns_still_returns_200_and_never_blocks():
    """Warnings inform; they do not become a 4xx (D3).

    The same rule the audit follows everywhere else in this router, asserted
    on the one route whose entire job is reporting problems — the most
    tempting place to turn a warning into a rejection.
    """
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    client.post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-17", "reason": "חופשה",
    })
    response = client.post("/api/schedule/check", json={
        "employee": "דנה", "shift_name": MORNING, "slot_date": "2026-08-17",
    })
    assert response.status_code == 200
    body = response.json()
    assert body["ok"] is False
    assert body["blocking"] is False
    assert body["reasons"], "a refusal-shaped answer with no reason given"


def test_a_check_that_warns_offers_alternatives():
    """A reason without a way out leaves the manager reading the grid."""
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    client.post("/api/schedule/constraints", json={
        "employee": "דנה", "constraint_date": "2026-08-17", "reason": "חופשה",
    })
    body = client.post("/api/schedule/check", json={
        "employee": "דנה", "shift_name": MORNING, "slot_date": "2026-08-17",
    }).json()
    alternatives = body["alternatives"]
    assert alternatives["employees"] or alternatives["slots"]


def test_checking_does_not_write():
    """`/check` is a read. Nothing it says lands until the manager acts."""
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    before = len(repo.changes)
    client.post("/api/schedule/check", json={
        "employee": "דנה", "shift_name": MORNING, "slot_date": "2026-08-17",
    })
    assert len(repo.changes) == before


# -- the board: the week containing a date ---------------------------------

def test_the_period_containing_a_date_is_served():
    """What the board opens on — the week the manager is actually in."""
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-16", "ends_on": "2026-08-22",
    })
    body = client.get("/api/schedule/at", params={"day": "2026-08-19"}).json()
    assert body is not None
    assert body["starts_on"] == "2026-08-16"


def test_a_date_outside_every_period_answers_null():
    """Null rather than 404: "no schedule that week" is a normal state the
    board renders as an empty week, not an error."""
    app, repo = _build_app([])
    client = _client(app)
    client.post("/api/schedule/blank", json={
        "starts_on": "2026-08-16", "ends_on": "2026-08-22",
    })
    assert client.get(
        "/api/schedule/at", params={"day": "2026-10-01"}
    ).json() is None


def test_a_member_reaches_only_published_periods_by_date():
    """The board is reachable from the read-only side, and a draft is still
    the manager's working state until they publish it (D5)."""
    app, repo = _build_app([])
    boss = _client(app)
    boss.post("/api/schedule/blank", json={
        "starts_on": "2026-08-16", "ends_on": "2026-08-22",
    })
    member = _client(app, role=ROLE_MEMBER)
    assert member.get(
        "/api/schedule/at", params={"day": "2026-08-19"}
    ).json() is None


def test_one_workspace_cannot_read_another_by_date():
    """D10: the team comes off the cookie, so another team's week is simply
    not there rather than forbidden."""
    app, repo = _build_app([])
    _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-16", "ends_on": "2026-08-22",
    })
    other = _client(app, team=OTHER_TEAM)
    assert other.get(
        "/api/schedule/at", params={"day": "2026-08-19"}
    ).json() is None


# -- a partial profile is a fork in the flow, not a dead end ---------------

def _ended_early(repo, **overrides):
    """A profile shaped like one `interview_service.end` wrote.

    The escape hatch is allowed to complete an interview with whatever was
    collected, recording the rest on `completeness`. That profile is real and
    `team_profile()` serves it, so every "is the interview done" check passes
    and only the shift vocabulary is actually absent.
    """
    profile = {
        "workplace": {"name": "מוקד"},
        "employees": [{"name": "דנה"}],
        "shifts": [],
        "rules": [],
        "completeness": {
            "complete": False,
            "missing_topics": ["לא הוגדר אף סוג משמרת."],
            "open_points": ["לא סוכם מי עובד בסופי שבוע."],
        },
    }
    profile.update(overrides)
    repo.profiles[TEAM] = profile
    return profile


def test_a_blank_period_over_a_profile_with_no_shifts_says_what_is_missing():
    """The bug this guard exists for.

    An interview ended early leaves a profile that exists but carries no
    shift vocabulary. It used to pass the `if not profile` check, reach
    `build_slots`, come back empty, and fail as an unexplained 502 -- the
    manual path looking broken rather than unfinished. The refusal now names
    the gap and says the interview can be resumed.
    """
    app, repo = _build_app([])
    _ended_early(repo)
    response = _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    assert response.status_code == 502
    body = response.json()
    assert "משמרות" in body["detail"]
    # The part a client can act on: which topics, and that there is a door.
    assert body["can_resume_interview"] is True
    assert "לא הוגדר אף סוג משמרת." in body["gaps"]
    assert body["blocks"], "a gap that blocks says what it costs"


def test_generating_over_a_profile_with_no_shifts_fails_the_same_way():
    """Both building paths ask one question, so they give one answer.

    They differ in whether a model runs; they do not differ in what a profile
    must contain before a grid exists, and a manager who hit the wall on one
    button must not find the other reporting something else.
    """
    app, repo = _build_app([])
    _ended_early(repo)
    response = _client(app).post("/api/schedule/generate", json={})
    assert response.status_code == 502
    body = response.json()
    assert body["can_resume_interview"] is True
    assert "לא הוגדר אף סוג משמרת." in body["gaps"]


def test_a_shift_the_grid_builder_would_skip_counts_as_no_vocabulary():
    """The gate applies the builder's own test, not a laxer one.

    `build_slots` needs a usable name. A nameless entry makes `shifts`
    non-empty while producing no rows, so a gate checking only for a
    non-empty list would pass it through to the empty grid this whole change
    exists to stop.
    """
    app, repo = _build_app([])
    _ended_early(repo, shifts=[{"start_time": "07:00", "end_time": "15:00"}])
    response = _client(app).post("/api/schedule/blank", json={})
    assert response.status_code == 502
    assert response.json()["can_resume_interview"] is True


def test_shifts_that_never_run_in_the_window_blame_the_dates_not_the_interview():
    """A different failure deserves a different sentence.

    The vocabulary is complete here -- the shift simply does not run on the
    days asked for. Re-opening the interview would teach nothing, so this
    stays an ordinary refusal with no resume offered.
    """
    app, repo = _build_app([])
    repo.profiles[TEAM] = dict(PROFILE, shifts=[{
        "name": MORNING, "start_time": "07:00", "end_time": "15:00",
        # Sunday only; the window below is Monday-Tuesday.
        "days": ["ראשון"], "is_on_call": False,
        "staffing": [{"days": [], "headcount": 1, "required_roles": []}],
    }])
    response = _client(app).post("/api/schedule/blank", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    })
    assert response.status_code == 502
    body = response.json()
    assert "תאריכים" in body["detail"]
    assert "can_resume_interview" not in body


# -- how wide a build's model calls are ------------------------------------


class _FakeSettings:
    """The runtime-settings store, reduced to the one field this reads."""

    def __init__(self, mode):
        self._mode = mode

    def get(self):
        return SimpleNamespace(schedule_generation_mode=self._mode)


def test_week_mode_builds_a_week_in_one_model_call():
    launcher = _DeferredLauncher()
    # A complete, legal week in one answer: every slot staffed, and the two
    # of them alternating so neither runs past the consecutive-days ceiling.
    # Nothing left for a repair call to chase, which is what makes "one model
    # call" the claim this test is actually making.
    week = [
        {
            "employee": "דנה" if day % 2 else "יוסי",
            "shift": MORNING,
            "date": "2026-08-%d" % day,
            "reason": "מאזן את הבקרים בין דנה ליוסי",
        }
        for day in range(17, 24)
    ]
    app, _ = _build_app(
        [_generation(week)],
        launch=launcher,
        settings=_FakeSettings("week"),
    )
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-23",
    }).json()

    # One span covering the week, but still seven days of progress: the bar
    # measures the period, not the number of calls.
    assert len(started["generation"]["days"]) == 1
    assert started["generation"]["total_days"] == 7
    assert started["generation"]["mode"] == "week"

    client.post("/api/schedule/generate/%s/run" % started["id"])
    launcher.run_next()

    body = client.get("/api/schedule/%s" % started["id"]).json()
    assert body["generation"]["status"] == "complete"
    assert body["generation"]["completed_days"] == 7
    assert len({row["date"] for row in body["assignments"]}) == 7


def test_day_mode_is_the_default_and_asks_once_per_date():
    launcher = _DeferredLauncher()
    app, _ = _build_app(
        [
            _generation([{
                "employee": "דנה", "shift": MORNING,
                "date": "2026-08-%d" % day, "reason": "דנה זמינה",
            }])
            for day in (17, 18)
        ],
        launch=launcher,
    )
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()

    assert started["generation"]["mode"] == "day"
    assert len(started["generation"]["days"]) == 2

    client.post("/api/schedule/generate/%s/run" % started["id"])
    launcher.run_next()
    assert client.get("/api/schedule/%s" % started["id"]).json()[
        "generation"
    ]["status"] == "complete"


def test_a_checkpoint_leaves_earlier_days_rows_alone():
    """Including their ids.

    Rewriting the whole period per day minted a fresh id for every row on
    every checkpoint, so an `assignment_id` the browser was holding — a drag
    opened mid-build, an employee's "what changed" row — pointed at nothing
    by the time it was used.
    """
    launcher = _DeferredLauncher()
    app, _ = _build_app(
        [
            _generation([{
                "employee": "דנה", "shift": MORNING,
                "date": "2026-08-%d" % day, "reason": "דנה זמינה",
            }])
            for day in (17, 18)
        ],
        launch=launcher,
    )
    client = _client(app)
    started = client.post("/api/schedule/generate/start", json={
        "starts_on": "2026-08-17", "ends_on": "2026-08-18",
    }).json()
    path = "/api/schedule/generate/%s/next" % started["id"]

    client.post(path)
    first = client.get("/api/schedule/%s" % started["id"]).json()
    monday = next(
        row for row in first["assignments"] if row["date"] == "2026-08-17"
    )

    client.post(path)
    second = client.get("/api/schedule/%s" % started["id"]).json()
    still = next(
        row for row in second["assignments"] if row["date"] == "2026-08-17"
    )

    assert still["id"] == monday["id"]
