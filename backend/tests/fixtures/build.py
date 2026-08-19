"""The two real layouts from `FILE_FORMATS.md`, rebuilt as `.xlsx`.

Built in code rather than committed as binaries so the structure that is
being tested is *readable* — the whole point of these two files is that they
disagree with each other structurally, and a checked-in `.xlsx` hides exactly
that. `BUILD_ORDER.md` step 6 says to build them first and make inference
pass both.

Sample A is shift-major and dense: dates across the top, shifts down the
side, one name per cell. Sample B is person-major and sparse, with a nested
`date x shift` header built from merged cells, and availability markers
living in the same grid as the assignments.
"""

from openpyxl import Workbook

MORNING = "בוקר"
EVENING = "צהריים"
ON_CALL = "כונן לילה"

# `לא זמינה` / `לא זמין` appear as cell values beside real names in Sample B.
UNAVAILABLE_F = "לא זמינה"
UNAVAILABLE_M = "לא זמין"


def sample_a():
    """Shift-major, dense, `d/M/yy` dates with a Hebrew weekday row beneath."""
    book = Workbook()
    sheet = book.active
    sheet.title = "סידור"
    sheet.sheet_view.rightToLeft = True

    dates = ["15/6/25", "16/6/25", "17/6/25", "18/6/25", "19/6/25"]
    weekdays = ["ראשון", "שני", "שלישי", "רביעי", "חמישי"]

    sheet.cell(row=1, column=1, value="משמרות")
    for index, date in enumerate(dates):
        sheet.cell(row=1, column=index + 2, value=date)
        sheet.cell(row=2, column=index + 2, value=weekdays[index])

    rows = [
        (MORNING, ["שובל ק.", "אלמוג ע.", "אורי א.", "אלמוג ע.", "תאיר ג."]),
        (EVENING, ["ניצן ש.", "ניצן ש.", "אלמוג ע.", "סאלי א.", "סאלי א."]),
    ]
    for offset, (shift_name, people) in enumerate(rows):
        row = 3 + offset
        sheet.cell(row=row, column=1, value=shift_name)
        for index, person in enumerate(people):
            sheet.cell(row=row, column=index + 2, value=person)
    return book


def sample_b():
    """Person-major, sparse, nested `date x shift` header from merged cells.

    The date row is merged across its three shift sub-columns, which is what
    makes the header nested rather than merely two-line: reading the date row
    literally gives a value in the first sub-column and blanks in the other
    two, and only the merge says those blanks belong to the date.
    """
    book = Workbook()
    sheet = book.active
    sheet.title = "סידור"
    sheet.sheet_view.rightToLeft = True

    dates = ["2.2 ראשון", "3.2 שני"]
    shifts = [MORNING, EVENING, ON_CALL]

    for index, date in enumerate(dates):
        first = 2 + index * len(shifts)
        sheet.cell(row=1, column=first, value=date)
        sheet.merge_cells(
            start_row=1, start_column=first,
            end_row=1, end_column=first + len(shifts) - 1,
        )
        for offset, shift_name in enumerate(shifts):
            sheet.cell(row=2, column=first + offset, value=shift_name)

    # One lane per person. Most cells are empty; a name appears where that
    # person is placed, and an availability marker appears in the same grid.
    lanes = [
        [(1, 1, "מאור נ")],
        [(0, 1, "נועה ר")],
        [],
        [(0, 1, "יערה ש")],
        [(1, 1, UNAVAILABLE_F)],
    ]
    for offset, placements in enumerate(lanes):
        row = 3 + offset
        for date_index, shift_index, value in placements:
            column = 2 + date_index * len(shifts) + shift_index
            sheet.cell(row=row, column=column, value=value)
    return book


__all__ = ["sample_a", "sample_b", "MORNING", "EVENING", "ON_CALL"]
