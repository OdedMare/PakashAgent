"""Exporting a period as `.xlsx`.

Pure functions over a stored schedule, so the whole contract is testable with
no database and no model. What is pinned here is the part that is not
cosmetic: the **layout**, which is Sample A from `FILE_FORMATS.md` because the
importer is being built to read that shape, and the distinction between a
shift that does not run and a shift nobody is on.
"""

import io

import pytest
from openpyxl import load_workbook

from app.bl.export import UNFILLED, as_workbook, filename
from app.common.errors import AgentError

MORNING = "בוקר"
EVENING = "צהריים"
ON_CALL = "כונן לילה"


def _schedule(slots=None, assignments=None, **overrides):
    schedule = {
        "id": "sched-1",
        "status": "draft",
        "starts_on": "2026-08-17",
        "ends_on": "2026-08-18",
        "slots": slots if slots is not None else [
            {"shift_name": MORNING, "slot_date": "2026-08-17"},
            {"shift_name": MORNING, "slot_date": "2026-08-18"},
        ],
        "assignments": assignments if assignments is not None else [
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
        ],
    }
    schedule.update(overrides)
    return schedule


def _sheet(schedule, title=""):
    return load_workbook(io.BytesIO(as_workbook(schedule, title=title))).active


def _grid(sheet):
    """`(row_label, column_date) -> cell value`, skipping the title rows."""
    dates = {}
    for column in range(2, sheet.max_column + 1):
        dates[column] = sheet.cell(row=4, column=column).value
    grid = {}
    for row in range(6, sheet.max_row + 1):
        label = sheet.cell(row=row, column=1).value
        if not label:
            continue
        for column, date in dates.items():
            grid[(label, date)] = sheet.cell(row=row, column=column).value
    return grid


# -- the layout ------------------------------------------------------------

def test_the_sheet_is_shift_major_with_dates_across_the_top():
    """Sample A's shape. Not a preference: the importer is being built to
    read it, so a week can leave, be edited, and come back."""
    sheet = _sheet(_schedule())

    assert sheet.cell(row=4, column=1).value == "משמרות"
    assert sheet.cell(row=4, column=2).value == "17.8.2026"
    assert sheet.cell(row=5, column=2).value == "יום שני"
    assert sheet.cell(row=6, column=1).value == MORNING


def test_the_sheet_is_right_to_left():
    """Hebrew in an LTR sheet puts the first column on the wrong side, which
    is not cosmetic for a grid whose columns are days."""
    assert _sheet(_schedule()).sheet_view.rightToLeft is True


def test_assigned_people_land_in_their_slot():
    grid = _grid(_sheet(_schedule()))

    assert grid[(MORNING, "17.8.2026")] == "דנה"


def test_several_people_on_one_slot_share_the_cell():
    sheet = _sheet(_schedule(assignments=[
        {"employee": "יוסי", "shift": MORNING, "date": "2026-08-17"},
        {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
    ]))

    # Sorted, so two identical schedules never export differently.
    assert _grid(sheet)[(MORNING, "17.8.2026")] == "דנה\nיוסי"


# -- the distinction that matters ------------------------------------------

def test_an_unstaffed_slot_is_named_rather_than_left_blank():
    """The case the manager most needs to notice. A blank cell reads as
    "nothing happens then"."""
    grid = _grid(_sheet(_schedule()))

    assert grid[(MORNING, "18.8.2026")] == UNFILLED


def test_a_shift_that_does_not_run_that_day_is_blank():
    """Not the same as unstaffed, and conflating them would report a gap
    that does not exist. Evening runs only on the first day here."""
    sheet = _sheet(_schedule(
        slots=[
            {"shift_name": MORNING, "slot_date": "2026-08-17"},
            {"shift_name": EVENING, "slot_date": "2026-08-17"},
            {"shift_name": MORNING, "slot_date": "2026-08-18"},
        ],
        assignments=[
            {"employee": "דנה", "shift": MORNING, "date": "2026-08-17"},
            {"employee": "רון", "shift": EVENING, "date": "2026-08-17"},
        ],
    ))
    grid = _grid(sheet)

    assert grid[(EVENING, "17.8.2026")] == "רון"
    # openpyxl reads an empty cell back as None, which is what a blank in a
    # spreadsheet actually is.
    assert not grid[(EVENING, "18.8.2026")]
    assert grid[(MORNING, "18.8.2026")] == UNFILLED


def test_the_grid_is_built_from_slots_not_only_assignments():
    """An entirely unstaffed period still exports its shape — walking the
    assignments alone would produce an empty sheet."""
    sheet = _sheet(_schedule(assignments=[]))

    assert _grid(sheet)[(MORNING, "17.8.2026")] == UNFILLED


def test_an_assignment_outside_the_grid_is_still_exported():
    """It is still real to the person standing there."""
    sheet = _sheet(_schedule(assignments=[
        {"employee": "דנה", "shift": ON_CALL, "date": "2026-08-17"},
    ]))

    assert _grid(sheet)[(ON_CALL, "17.8.2026")] == "דנה"


# -- the rest --------------------------------------------------------------

def test_the_workplace_name_titles_the_sheet():
    sheet = _sheet(_schedule(), title="מוקד")

    assert sheet.cell(row=1, column=1).value == "מוקד"
    assert sheet.cell(row=2, column=1).value == "17.8.2026 – 18.8.2026"


def test_an_empty_schedule_refuses_rather_than_exporting_a_blank_sheet():
    with pytest.raises(AgentError):
        as_workbook({"slots": [], "assignments": []})


def test_the_filename_carries_the_period_and_is_ascii():
    """A Hebrew filename is correct but travels badly through
    `Content-Disposition`."""
    name = filename(_schedule(), "xlsx")

    assert name == "pakash_2026-08-17_2026-08-18.xlsx"
    name.encode("ascii")
