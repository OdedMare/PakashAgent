"""Handing a finished schedule back as a file.

The other half of [D6](../../../docs/DECISIONS.md#d6--the-boss-can-author-or-generate):
the product absorbs a schedule the manager already had, so it must also give
one back. The share link is a fine way for the team to *read* the roster, but
a small workplace keeps a copy in a spreadsheet, and a product that can only
be read inside its own tab is one step removed from how the work circulates.

**One format, and it is not a preference.** The sheet is laid out shift-major
with dates across the top — the shape of Sample A in
[`FILE_FORMATS.md`](../../../docs/FILE_FORMATS.md). Exporting in the layout
the importer is being built to read means a week can leave, be edited in
Excel, and come back. A prettier bespoke layout would be a file this product
cannot read.

**A message for the team is not here.** Posting the week to a group chat is a
thing the manager asks the *agent* for, in the conversation they are already
having ([D17](../../../docs/DECISIONS.md#d17--a-schedule-leaves-as-a-file-a-message-is-something-the-agent-writes)) —
it is writing, and writing in this product is the agent's job. A fixed
template here would be a second voice saying the same thing worse.

**Pure functions over a stored schedule. No model call and no repository.**
Nothing here decides anything — it re-presents what the manager already
confirmed.

Hebrew is data here, not presentation: weekday names and shift names come
from the workplace's own vocabulary
([D9](../../../docs/DECISIONS.md#d9--shift-vocabulary-is-per-workplace)).
"""

import datetime
import io
from typing import Any, Dict, List, Optional

from app.common.errors import AgentError

# Matching `scheduler._HEBREW_WEEKDAYS`: Monday-first, because that is what
# `date.weekday()` returns. Duplicated rather than imported to keep this
# module free of the scheduler -- it is a table of Hebrew names, and the two
# would have to change together anyway if the language ever did.
_HEBREW_WEEKDAYS = (
    "יום שני", "יום שלישי", "יום רביעי", "יום חמישי",
    "יום שישי", "שבת", "יום ראשון",
)

# A slot nobody is on. Said out loud rather than left blank: an empty cell in
# a group chat reads as "nothing that day", and an unstaffed shift is the one
# thing the manager most needs somebody to notice.
UNFILLED = "— לא מאויש —"


def as_workbook(schedule: dict, title: str = "") -> bytes:
    """The week as `.xlsx`, shift-major with dates across the top.

    Deliberately the shape of Sample A in `FILE_FORMATS.md` rather than a
    layout of our own: the importer is being built to read that shape, so a
    week exported here can be edited in Excel and brought back. A prettier
    bespoke layout would be a file this product cannot read.

    Right-to-left is set on the sheet itself. Hebrew in a left-to-right sheet
    renders with the first column on the wrong side, which is not cosmetic
    for a grid whose columns are days.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:  # pragma: no cover - openpyxl is a hard dependency
        raise AgentError("ייצוא לאקסל אינו זמין בשרת הזה")

    days = _by_day(schedule)
    if not days:
        raise AgentError("אין מה לייצא: הסידור ריק")

    dates = sorted(days)
    shifts = _ordered_shifts_across(days)

    book = Workbook()
    sheet = book.active
    sheet.title = "סידור"
    sheet.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="EEECE3")
    thin = Side(style="thin", color="D9D5CA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    heading = (title or "").strip() or "סידור עבודה"
    sheet.cell(row=1, column=1, value=heading).font = Font(bold=True, size=13)
    period = _period(schedule)
    if period:
        sheet.cell(row=2, column=1, value=period)

    # Two header rows, matching Sample A: the date, then its weekday beneath.
    top = 4
    label = sheet.cell(row=top, column=1, value="משמרות")
    label.font = Font(bold=True)
    label.fill = header_fill
    label.border = border
    label.alignment = centered
    sheet.cell(row=top + 1, column=1).fill = header_fill
    sheet.cell(row=top + 1, column=1).border = border

    for index, date in enumerate(dates):
        column = index + 2
        head = sheet.cell(row=top, column=column, value=_human_date(date))
        day = sheet.cell(row=top + 1, column=column, value=_weekday(date))
        for cell in (head, day):
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.border = border
            cell.alignment = centered
        sheet.column_dimensions[get_column_letter(column)].width = 16

    sheet.column_dimensions["A"].width = 18

    for offset, shift_name in enumerate(shifts):
        row = top + 2 + offset
        name = sheet.cell(row=row, column=1, value=shift_name)
        name.font = Font(bold=True)
        name.fill = header_fill
        name.border = border
        name.alignment = centered
        lines = 1
        for index, date in enumerate(dates):
            people = days[date].get(shift_name)
            # `None` means this shift does not run that day, which is not the
            # same as running with nobody on it. An empty cell says the
            # former; `UNFILLED` says the latter, and conflating them would
            # report a gap that does not exist.
            if people is None:
                value = ""
            else:
                value = "\n".join(people) if people else UNFILLED
                lines = max(lines, len(people) or 1)
            cell = sheet.cell(row=row, column=index + 2, value=value)
            cell.border = border
            cell.alignment = centered
        # Sized to the fullest cell in the row rather than fixed. A shift run
        # by four or ten people writes four or ten names into one cell, and a
        # fixed height silently crops all but the first two -- the exported
        # week would then show fewer people than the schedule holds, which is
        # worse than an ugly file. Read back it is unaffected either way:
        # `importer._split_names` splits the cell on its newlines.
        sheet.row_dimensions[row].height = max(34, 15 * lines + 8)

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def filename(schedule: dict, extension: str) -> str:
    """A filename carrying the period, so a folder of these stays readable.

    ASCII only. A Hebrew filename is correct but travels badly through
    `Content-Disposition`, and the period is what actually distinguishes one
    export from the next.
    """
    starts = _iso(schedule.get("starts_on")) or "schedule"
    ends = _iso(schedule.get("ends_on"))
    stem = "%s_%s" % (starts, ends) if ends and ends != starts else starts
    return "pakash_%s.%s" % (stem, extension)


def _by_day(schedule: Optional[dict]) -> Dict[str, Dict[str, List[str]]]:
    """`date -> shift -> the people on it`.

    Built from the **slot grid** and then filled from the assignments, not
    from the assignments alone. An unstaffed shift leaves no assignment row,
    so walking only those would drop exactly the slot worth reporting -- the
    same trap `audit.py` documents for the unfilled warning.
    """
    schedule = schedule if isinstance(schedule, dict) else {}
    days: Dict[str, Dict[str, List[str]]] = {}

    for slot in schedule.get("slots") or []:
        if not isinstance(slot, dict):
            continue
        date = _iso(slot.get("slot_date"))
        shift_name = _text(slot.get("shift_name"))
        if not date or not shift_name:
            continue
        days.setdefault(date, {}).setdefault(shift_name, [])

    for row in schedule.get("assignments") or []:
        if not isinstance(row, dict):
            continue
        date = _iso(row.get("date"))
        shift_name = _text(row.get("shift"))
        employee = _text(row.get("employee"))
        if not date or not employee:
            continue
        # An assignment naming a slot the grid does not have is still real to
        # the person standing there, so it is kept rather than dropped.
        days.setdefault(date, {}).setdefault(shift_name, []).append(employee)

    for shifts in days.values():
        for people in shifts.values():
            people.sort()
    return days


def _ordered_shifts_across(
    days: Dict[str, Dict[str, List[str]]]
) -> List[str]:
    """Every shift name appearing anywhere in the period, as rows.

    One row per name across the whole week rather than per day: the grid is
    rectangular, and a shift that runs only on some days simply leaves those
    other cells empty.
    """
    names = set()
    for shifts in days.values():
        names.update(shifts)
    return sorted(names)


def _period(schedule: dict) -> str:
    starts = _iso(schedule.get("starts_on"))
    ends = _iso(schedule.get("ends_on"))
    if not starts:
        return ""
    if not ends or ends == starts:
        return _human_date(starts)
    return "%s – %s" % (_human_date(starts), _human_date(ends))


def _weekday(iso: str) -> str:
    day = _parse(iso)
    return _HEBREW_WEEKDAYS[day.weekday()] if day else ""


def _human_date(iso: str) -> str:
    """`d.M.yyyy` — how the source files write dates, not ISO."""
    day = _parse(iso)
    if day is None:
        return iso
    return "%d.%d.%d" % (day.day, day.month, day.year)


def _parse(iso: str) -> Optional[datetime.date]:
    try:
        return datetime.datetime.strptime(iso[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _iso(value: Any) -> str:
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return _text(value)[:10]


def _text(value: Any) -> str:
    return value.strip() if isinstance(value, str) else ""


__all__ = ["as_workbook", "filename", "UNFILLED"]
