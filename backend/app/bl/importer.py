"""Reading a schedule the workplace already had, without a template.

[D7](../../../docs/DECISIONS.md#d7--import-infers-layout-boss-confirms): there
is no fixed template. The two real files in
[`FILE_FORMATS.md`](../../../docs/FILE_FORMATS.md) are structurally different
from each other — one is shift-major with dates across the top, the other is
person-major with a nested `date x shift` header — so an importer built for
either fails on the other. What is inferred here is therefore not orientation
but **axis semantics**: which axis carries time, whether shift is nested under
date or is an axis of its own, and whether the non-time lanes are shifts or
people.

**Nothing here persists anything.** Inference produces an interpretation the
manager reads and approves, and only then does `schedule_service` write. That
ordering is the whole of D7, and it is why this module is handed no
repository.

**No model call either.** Layout inference is grid arithmetic — counting how
many cells parse as dates, how many header labels match the declared shift
vocabulary — and code that counts cannot hallucinate a person into a shift
they did not work. The model's judgment enters one layer up, in
`learn.py`, where prose is actually being read. The line is
[D3](../../../docs/DECISIONS.md#d3--the-agent-decides-code-only-audits-)'s,
applied to reading rather than to deciding.

Shift names are matched against the vocabulary the interview collected
([D9](../../../docs/DECISIONS.md#d9--shift-vocabulary-is-per-workplace))
rather than inferred from scratch. Hebrew is data here: weekday names,
availability markers and two date formats (`d/M/yy` and `d.M`, the latter
with no year at all) are values to parse, not text to display.
"""

import datetime
import re
from typing import Any, Dict, List, Optional, Tuple

from app.common.errors import AgentError

# A sheet bigger than this is not a shift schedule. Bounded because every
# cell is visited several times during inference.
_MAX_ROWS = 400
_MAX_COLUMNS = 200
_MAX_FILE_BYTES = 20 * 1024 * 1024

# How far a header row may sit from the top before we stop looking for it.
# Real files carry a title and a blank line above the grid -- `export.py`
# writes exactly that -- but a header ten rows down means the sheet is not
# the shape we think it is.
_MAX_HEADER_SCAN = 10

# Hebrew weekday names as the source files write them, with and without the
# `יום ` prefix. Data, not presentation: they appear *inside* header cells
# ("2.2 ראשון") and are what disambiguates a date column from a stray number.
_HEBREW_WEEKDAYS = (
    "ראשון", "שני", "שלישי", "רביעי", "חמישי", "שישי", "שבת",
)

# Availability markers living in the assignment grid (Sample B). Matched as
# whole cell values after normalisation, never as substrings: a person named
# with one of these words inside a longer sentence is not a marker.
_UNAVAILABLE = ("לא זמין", "לא זמינה", "לא יכול", "לא יכולה", "חופש", "מחלה")

# The row-header label Sample A puts above the shift column. Its presence is
# evidence for shift-major, never a requirement -- a file without it is still
# read, just with one less signal.
_SHIFT_HEADER_HINTS = ("משמרות", "משמרת")


class Interpretation(object):
    """What the importer believes a file says, before anyone approves it.

    Deliberately a plain object with a `to_dict`: it crosses the HTTP
    boundary to be confirmed, and it is the only thing the manager sees
    before a write happens.
    """

    def __init__(
        self,
        layout: str,
        shifts: List[str],
        people: List[str],
        dates: List[str],
        assignments: List[dict],
        unavailability: List[dict],
        warnings: Optional[List[str]] = None,
    ):
        self.layout = layout
        self.shifts = shifts
        self.people = people
        self.dates = dates
        self.assignments = assignments
        self.unavailability = unavailability
        self.warnings = warnings or []

    @property
    def starts_on(self) -> str:
        return self.dates[0] if self.dates else ""

    @property
    def ends_on(self) -> str:
        return self.dates[-1] if self.dates else ""

    def summary(self) -> str:
        """The one sentence D7 asks for, in Hebrew, for the confirm screen."""
        parts = ["%d אנשים" % len(self.people)]
        if self.dates:
            parts.append("%s עד %s" % (
                _human(self.starts_on), _human(self.ends_on),
            ))
        parts.append("%d משמרות" % len(self.shifts))
        parts.append("%d שיבוצים" % len(self.assignments))
        if self.unavailability:
            parts.append("%d סימוני אי-זמינות" % len(self.unavailability))
        return ", ".join(parts)

    def to_dict(self) -> dict:
        return {
            "layout": self.layout,
            "shifts": self.shifts,
            "people": self.people,
            "dates": self.dates,
            "starts_on": self.starts_on,
            "ends_on": self.ends_on,
            "assignments": self.assignments,
            "unavailability": self.unavailability,
            "warnings": self.warnings,
            "summary": self.summary(),
        }


def read_grid(data: bytes, filename: str = "") -> List[List[str]]:
    """A file's first sheet (or its tables) as a rectangular grid of strings.

    Merged cells are **unmerged by value**: openpyxl reports a merged range as
    one value plus blanks, and Sample B's nested header depends entirely on
    those blanks belonging to the date above them. Filling them in here is
    what lets inference treat the header as a real `date x shift` grid rather
    than a row with holes in it.
    """
    return read_grids(data, filename)[0][1]


def read_grids(
    data: bytes, filename: str = ""
) -> List[Tuple[str, List[List[str]]]]:
    """Every usable worksheet/table in one upload.

    A workbook often opens on a summary or instructions tab while the real
    schedules live in the other tabs. Returning every sheet lets the caller
    infer each one independently; non-schedule tabs then become ordinary
    per-sheet failures instead of hiding the schedules behind them.
    """
    if not data:
        raise AgentError("הקובץ ריק")
    if len(data) > _MAX_FILE_BYTES:
        raise AgentError("הקובץ גדול מדי (עד 20MB לקובץ)")

    name = (filename or "").lower()
    if name.endswith(".docx"):
        return [("", _read_docx(data))]
    if name.endswith(".xls"):
        raise AgentError(
            "פורמט Excel הישן (.xls) אינו נתמך; "
            "שמור את הקובץ כ־.xlsx ונסה שוב"
        )
    return _read_xlsx_sheets(data)


def infer(grid: List[List[str]], profile: Optional[dict] = None) -> Interpretation:
    """Read a grid as a schedule, inferring what its axes mean.

    Tries both layouts and keeps whichever explains more of the sheet. The
    two are not variants of one parser: shift-major reads lanes as shifts and
    cells as people, person-major reads lanes as people and cells as
    placements, and the header shapes differ too. Scoring them against each
    other is more honest than a heuristic that picks first and hopes.
    """
    grid = _bounded(grid)
    if not grid:
        raise AgentError("הקובץ ריק או לא נקרא")

    vocabulary = _vocabulary(profile)
    candidates = []
    for reader in (_read_person_major, _read_shift_major, _read_dateonly):
        try:
            found = reader(grid, vocabulary)
        except AgentError:
            found = None
        if found is not None:
            candidates.append(found)

    if not candidates:
        raise AgentError(
            "לא הצלחתי לזהות את מבנה הקובץ. "
            "ודא שיש בו שורת תאריכים ושמות משמרות"
        )
    # More explained cells wins, but a layout that found a real shift axis
    # always beats one that did not: `_read_dateonly` explains every filled
    # cell by construction, so scoring it against the others would let it
    # win whenever it applies and flatten a shift axis that was really
    # there. It is a fallback, and fallbacks go last.
    structured = [
        found for found in candidates if found[1].layout != "date_only"
    ]
    ranked = sorted(
        structured or candidates, key=lambda found: found[0], reverse=True
    )
    best = ranked[0]
    result = best[1]
    if (
        len(ranked) > 1
        and ranked[1][1].layout != result.layout
        and ranked[1][0] >= best[0] * 0.85
    ):
        result.warnings.append(
            "המבנה אינו חד־משמעי; בדוק שהשורות והעמודות "
            "פורשו נכון לפני האישור"
        )
    if _has_yearless_dates(grid):
        result.warnings.append(
            "בקובץ יש תאריכים ללא שנה; השנה הנוכחית הונחה — "
            "ודא אותה לפני האישור"
        )
    return _deduplicated(result)


# -- xlsx / docx reading ---------------------------------------------------

def _read_xlsx(data: bytes) -> List[List[str]]:
    return _read_xlsx_sheets(data)[0][1]


def _read_xlsx_sheets(data: bytes) -> List[Tuple[str, List[List[str]]]]:
    try:
        import io

        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - openpyxl is a hard dependency
        raise AgentError("קריאת קבצי אקסל אינה זמינה בשרת הזה")

    try:
        book = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:
        raise AgentError("לא הצלחתי לפתוח את קובץ האקסל")

    sheets = [
        sheet for sheet in book.worksheets
        if sheet.sheet_state == "visible"
    ] or list(book.worksheets)
    found = []
    for sheet in sheets:
        grid = _sheet_grid(sheet)
        if any(any(cell for cell in row) for row in grid):
            found.append((sheet.title, grid))
    if not found:
        raise AgentError("קובץ האקסל ריק")
    return found


def _sheet_grid(sheet) -> List[List[str]]:
    """One worksheet, bounded before iteration rather than afterwards.

    Excel files frequently carry formatting down to row 1,048,576. Asking
    openpyxl to iterate the reported dimensions before `_bounded` runs makes
    a visually tiny sheet need huge time and memory, so the limits belong at
    the read boundary.
    """
    grid = [
        [_text(value) for value in row]
        for row in sheet.iter_rows(
            min_row=1,
            max_row=min(sheet.max_row, _MAX_ROWS),
            min_col=1,
            max_col=min(sheet.max_column, _MAX_COLUMNS),
            values_only=True,
        )
    ]

    # Merged ranges carry their value only in the top-left cell. Sample B's
    # date header is merged across its shift sub-columns, so without this the
    # nested header reads as a date followed by blanks.
    for merged in sheet.merged_cells.ranges:
        top, left = merged.min_row - 1, merged.min_col - 1
        if top >= len(grid) or left >= len(grid[top]):
            continue
        value = grid[top][left]
        if not value:
            continue
        for row in range(
            merged.min_row - 1, min(merged.max_row, _MAX_ROWS)
        ):
            for column in range(
                merged.min_col - 1, min(merged.max_col, _MAX_COLUMNS)
            ):
                if row < len(grid) and column < len(grid[row]):
                    grid[row][column] = value
    return _rectangular(grid)


def _read_docx(data: bytes) -> List[List[str]]:
    """The first table in a Word document, as a grid.

    A schedule pasted into Word is still a table, and Word writes a merged
    cell by repeating its content in each covered cell -- which is the
    behaviour the nested header needs, for free.

    Unlike the two `.xlsx` samples this has no real fixture behind it, so it
    is kept to the narrowest thing that can work: find a table, read its
    cells, hand the same grid to the same inference.
    """
    try:
        import io
        import zipfile
        from xml.etree import ElementTree
    except ImportError:  # pragma: no cover - stdlib
        raise AgentError("קריאת קבצי וורד אינה זמינה בשרת הזה")

    namespace = (
        "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    )
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as archive:
            xml = archive.read("word/document.xml")
    except Exception:
        raise AgentError("לא הצלחתי לפתוח את קובץ הוורד")

    try:
        root = ElementTree.fromstring(xml)
    except ElementTree.ParseError:
        raise AgentError("לא הצלחתי לפתוח את קובץ הוורד")

    table = root.iter(namespace + "tbl")
    for element in table:
        grid = []
        for row in element.iter(namespace + "tr"):
            cells = []
            for cell in row.iter(namespace + "tc"):
                text = "".join(
                    node.text or "" for node in cell.iter(namespace + "t")
                )
                cells.append(_text(text))
            grid.append(cells)
        if grid:
            return _rectangular(grid)
    raise AgentError("לא נמצאה טבלה במסמך")


# -- layout readers --------------------------------------------------------

def _read_shift_major(
    grid: List[List[str]], vocabulary: List[str]
) -> Optional[Tuple[int, Interpretation]]:
    """Sample A: dates across the top, shifts down the side, people in cells.

    The date row is found by counting parseable dates rather than by position,
    because real files carry a title and a blank line above the grid.
    """
    header = _find_date_row(grid)
    if header is None:
        return None
    row_index, dates = header

    # A weekday row directly beneath the dates is part of the header, not
    # data. Recognised by content, not by offset.
    first_data = row_index + 1
    if first_data < len(grid) and _is_weekday_row(grid[first_data], dates):
        first_data += 1

    shifts: List[str] = []
    assignments: List[dict] = []
    unavailability: List[dict] = []
    people: List[str] = []
    explained = 0

    for row in grid[first_data:_MAX_ROWS]:
        label = _lane_label(row, [column for column, _iso in dates])
        if not label or label in _SHIFT_HEADER_HINTS:
            continue
        shift_name = _match_shift(label, vocabulary)
        if not shift_name:
            continue
        # A row whose *cells* are themselves shift names is a header, not a
        # lane of people -- which is exactly what Sample B's second row looks
        # like to this reader. Reading it as data invents a person per shift
        # name, so the row is skipped rather than scored.
        if _mostly_shifts(row, dates, vocabulary):
            continue
        if shift_name not in shifts:
            shifts.append(shift_name)
        for column, iso in dates:
            if column >= len(row):
                continue
            value = row[column]
            if not value:
                continue
            explained += 1
            for name in _split_names(value):
                if _is_unavailable(name):
                    unavailability.append({
                        "employee": "", "date": iso,
                        "shift": shift_name, "reason": name,
                    })
                    continue
                assignments.append({
                    "employee": name, "shift": shift_name, "date": iso,
                })
                if name not in people:
                    people.append(name)

    if not shifts or not assignments:
        return None
    notes = _warnings(dates)
    if unavailability:
        notes.append(
            "%d סימוני אי־זמינות אינם משויכים לאדם — "
            "השלם את השם לפני האישור" % len(unavailability)
        )
    return explained, Interpretation(
        layout="shift_major",
        shifts=shifts,
        people=sorted(people),
        dates=[iso for _column, iso in dates],
        assignments=assignments,
        unavailability=unavailability,
        warnings=notes,
    )


def _read_person_major(
    grid: List[List[str]], vocabulary: List[str]
) -> Optional[Tuple[int, Interpretation]]:
    """Sample B: a nested `date x shift` header, one lane per person.

    Requires two stacked header rows — dates above, shift names beneath,
    each date spanning its shifts — which is a strong enough shape that a
    sheet matching it is almost certainly this layout.

    Availability markers share the grid with assignments here, so every cell
    is *classified* rather than read: a marker becomes an unavailability row
    against the person whose lane it sits in, and a name becomes an
    assignment. That coexistence is the thing `FILE_FORMATS.md` warns about.
    """
    header = _find_nested_header(grid, vocabulary)
    if header is None:
        return None
    shift_row, columns = header

    shifts: List[str] = []
    for _column, _iso, shift_name in columns:
        if shift_name not in shifts:
            shifts.append(shift_name)

    assignments: List[dict] = []
    unavailability: List[dict] = []
    people: List[str] = []
    explained = 0

    data_columns = [column for column, _iso, _shift in columns]
    for row in grid[shift_row + 1:_MAX_ROWS]:
        # A lane's person comes from a label column where the sheet has one.
        # Sample B has none -- the name is written at the point of assignment
        # -- so the lane is identified from the names inside the row instead.
        # Taking the first non-empty cell would be wrong for exactly the row
        # that matters: a lane holding only `לא זמינה` would be attributed to
        # a person called `לא זמינה`.
        lane = _lane_label(row, data_columns)
        if not lane:
            lane = _lane_from_cells(row, data_columns)
        for column, iso, shift_name in columns:
            if column >= len(row):
                continue
            value = row[column]
            if not value:
                continue
            explained += 1
            for name in _split_names(value):
                if _is_unavailable(name):
                    # The marker names nobody; the lane does. A lane that
                    # holds *only* markers names nobody either -- in the real
                    # Sample B that person is identified by their placements
                    # elsewhere in the sheet, and a lane with none is a row
                    # this file simply does not say the owner of. It is
                    # recorded with an empty employee and warned about rather
                    # than dropped: an unattributed constraint is still a
                    # fact about the period, and the confirm screen is where
                    # a person can supply the missing name (D7).
                    unavailability.append({
                        "employee": lane, "date": iso,
                        "shift": shift_name, "reason": name,
                    })
                    continue
                assignments.append({
                    "employee": name, "shift": shift_name, "date": iso,
                })
                if name not in people:
                    people.append(name)
        if lane and lane not in people:
            people.append(lane)

    if not assignments and not unavailability:
        return None
    dates: List[str] = []
    for _column, iso, _shift in columns:
        if iso not in dates:
            dates.append(iso)

    notes = _warnings([(0, iso) for iso in dates])
    orphans = len([row for row in unavailability if not row["employee"]])
    if orphans:
        notes.append(
            "%d סימוני אי-זמינות בקובץ אינם משויכים לאדם — "
            "השלם את השם לפני האישור" % orphans
        )
    return explained, Interpretation(
        layout="person_major",
        shifts=shifts,
        people=sorted(people),
        dates=sorted(dates),
        assignments=assignments,
        unavailability=unavailability,
        warnings=notes,
    )


def _read_dateonly(
    grid: List[List[str]], vocabulary: List[dict]
) -> Optional[Tuple[int, Interpretation]]:
    """Dates and the people under them, with no shift axis at all.

    The plainest thing a manager keeps: a column per day, names beneath it,
    and nothing naming a shift anywhere -- because the workplace runs one
    shift, or because whoever wrote the sheet did not think it needed saying.
    Refusing this would mean the product only reads files that already look
    like its own export, which is the opposite of D7.

    Tried last, and it claims a sheet only when no row carries a lane label
    at all. A file that *does* label its rows is one of the other two
    layouts, and reading it here would flatten a real shift axis into
    nothing.

    The shift is left **empty** rather than named. There is genuinely no
    shift information in the file, and inventing a name -- `"בוקר"`, or
    anything else -- would be exactly the hardcoding D9 forbids. The confirm
    screen is where the manager says which shift this was, and an empty name
    is what makes the question visible instead of silently answered.
    """
    header = _find_date_row(grid)
    if header is None:
        return None
    row_index, dates = header

    first_data = row_index + 1
    if first_data < len(grid) and _is_weekday_row(grid[first_data], dates):
        first_data += 1

    columns = [column for column, _iso in dates]
    assignments: List[dict] = []
    unavailability: List[dict] = []
    people: List[str] = []
    explained = 0

    for row in grid[first_data:_MAX_ROWS]:
        # Any labelled row means this sheet has an axis this reader would
        # discard, so it is not ours to read.
        if _lane_label(row, columns):
            return None
        for column, iso in dates:
            if column >= len(row) or not row[column]:
                continue
            explained += 1
            for name in _split_names(row[column]):
                if _is_unavailable(name):
                    unavailability.append({
                        "employee": "", "date": iso,
                        "shift": "", "reason": name,
                    })
                    continue
                assignments.append({
                    "employee": name, "shift": "", "date": iso,
                })
                if name not in people:
                    people.append(name)

    if not assignments:
        return None
    notes = _warnings(dates)
    notes.append(
        "בקובץ אין שמות משמרות — כל השיבוצים יובאו ללא משמרת. "
        "בחר את המשמרת לפני האישור"
    )
    return explained, Interpretation(
        layout="date_only",
        shifts=[],
        people=sorted(people),
        dates=[iso for _column, iso in dates],
        assignments=assignments,
        unavailability=unavailability,
        warnings=notes,
    )


# -- header inference ------------------------------------------------------

def _find_date_row(grid: List[List[str]]) -> Optional[Tuple[int, List[tuple]]]:
    """The row carrying the most parseable dates, near the top.

    Counting rather than assuming row 1: `export.py` writes a title and a
    period line above its header, and the real files do the same.
    """
    best = None
    for index, row in enumerate(grid[:_MAX_HEADER_SCAN]):
        dates = []
        for column, value in enumerate(row):
            iso = _parse_date(value)
            if iso:
                dates.append((column, iso))
        # One date is enough. A sheet covering a single day is a real thing
        # a manager keeps, and requiring two would refuse it for being
        # short. Ambiguity is handled by *preferring* the row with more
        # dates, not by rejecting thin ones.
        if dates and (best is None or len(dates) > len(best[1])):
            best = (index, dates)
    return best


def _find_nested_header(
    grid: List[List[str]], vocabulary: List[str]
) -> Optional[Tuple[int, List[tuple]]]:
    """A date row directly above a shift row, dates spanning their shifts.

    Returns the shift row's index and one entry per column: `(column, date,
    shift)`. The date row's blanks have already been filled from the merged
    ranges by `read_grid`, which is what makes the pairing a straight
    column-by-column read rather than a span calculation.
    """
    found = _find_date_row(grid)
    if found is None:
        return None
    date_row, dates = found
    shift_row = date_row + 1
    if shift_row >= len(grid):
        return None

    by_column = dict(dates)
    columns = []
    for column, value in enumerate(grid[shift_row]):
        shift_name = _match_shift(value, vocabulary)
        if not shift_name:
            continue
        iso = by_column.get(column) or _carried_date(by_column, column)
        if iso:
            columns.append((column, iso, shift_name))

    # Nested means several shift columns under one date. Without that this is
    # Sample A's two-row header (dates then weekdays), which the shift-major
    # reader owns.
    per_date: Dict[str, int] = {}
    for _column, iso, _shift in columns:
        per_date[iso] = per_date.get(iso, 0) + 1
    if not columns or max(per_date.values()) < 2:
        return None
    return shift_row, columns


def _carried_date(by_column: Dict[int, str], column: int) -> str:
    """The nearest date at or left of a column.

    A fallback for a merged header openpyxl did not report as a range —
    some writers emit the spanned cells as genuinely empty instead.
    """
    for candidate in range(column, -1, -1):
        if candidate in by_column:
            return by_column[candidate]
    return ""


# -- cell classification ---------------------------------------------------

def _match_shift(value: str, vocabulary: List[dict]) -> str:
    """A header cell read as a shift name.

    The declared vocabulary
    ([D9](../../../docs/DECISIONS.md#d9--shift-vocabulary-is-per-workplace))
    is what a header is matched *against*, and a match returns the declared
    spelling -- so a sheet writing `בוקר `, `משמרת בוקר`, or the shift's own
    hours all land on the one name the workplace uses. Matching against
    declared names rather than inferring from scratch is D9's point, and it
    is still what happens first.

    **An unmatched header is not discarded.** The file is a real schedule the
    workplace ran, and a sheet naming a shift the interview never mentioned
    -- an old name, a one-off, or a column headed only by its hours -- is
    describing something that genuinely happened. Dropping it would silently
    lose real history and would make the product refuse exactly the old files
    it exists to absorb (D7). So the cell's own text is taken instead, and
    the confirm screen is where the manager reconciles it.

    This is not the hardcoding D9 forbids: nothing is *invented* here. A name
    is either the workplace's declared one or the one written in the
    manager's own file. What never happens is guessing a name from neither.
    """
    text = _normalise(value)
    if not text:
        return ""
    # A date or a weekday heads a different axis, and is never a shift.
    if _parse_date(value) or text in _HEBREW_WEEKDAYS:
        return ""

    # Declared name, or the declared shift whose hours these are: a column
    # headed `07:00-15:00` is the same shift as the one running those hours,
    # so it folds into it rather than being stored again under a different
    # label -- otherwise one shift appears twice in the vocabulary and every
    # hours-per-person tally splits across the two.
    declared = _declared_shift(value, vocabulary)
    if declared:
        return declared

    # Nothing matched. The sheet's own label stands for itself.
    return value.strip()


_HOURS = re.compile(
    r"^(\d{1,2})[:.](\d{2})\s*(?:-|\u2013|\u2014|עד)\s*(\d{1,2})[:.](\d{2})$"
)


def _parse_hours(value: str) -> str:
    """`07:00-15:00` normalised, or empty when the cell is not a time range.

    Normalised rather than compared literally because the same range gets
    written `7:00-15:00`, `07.00-15.00` and `07:00 - 15:00` across files by
    the same person in the same year.
    """
    match = _HOURS.match(_normalise(value))
    if not match:
        return ""
    start_h, start_m, end_h, end_m = (int(part) for part in match.groups())
    if start_h > 23 or end_h > 23 or start_m > 59 or end_m > 59:
        return ""
    return "%02d:%02d-%02d:%02d" % (start_h, start_m, end_h, end_m)


def _is_unavailable(value: str) -> bool:
    text = _normalise(value)
    return any(text == marker or text.startswith(marker) for marker in _UNAVAILABLE)


def _split_names(value: str) -> List[str]:
    """One cell into the people it names.

    `export.py` writes several people into a cell newline-separated, and the
    manager's own files use commas or slashes. The unfilled marker it writes
    is not a person.
    """
    from app.bl.export import UNFILLED

    parts = re.split(r"[\n,/;]+", value or "")
    names = []
    for part in parts:
        name = part.strip()
        if not name or name == UNFILLED or name.startswith("—"):
            continue
        names.append(name)
    return names


def _is_weekday_row(row: List[str], dates: List[tuple]) -> bool:
    """Whether a row under the dates is the weekday line, not data."""
    hits = 0
    for column, _iso in dates:
        if column < len(row) and _normalise(row[column]) in _HEBREW_WEEKDAYS:
            hits += 1
    return hits >= max(1, len(dates) // 2)


# -- dates -----------------------------------------------------------------

_DATE_PATTERNS = (
    ("%Y-%m-%d", r"^(\d{4})-(\d{1,2})-(\d{1,2})$"),
    ("%d/%m/%Y", r"^(\d{1,2})/(\d{1,2})/(\d{4})$"),
    ("%d/%m/%y", r"^(\d{1,2})/(\d{1,2})/(\d{2})$"),
    ("%d.%m.%Y", r"^(\d{1,2})\.(\d{1,2})\.(\d{4})$"),
    ("%d.%m.%y", r"^(\d{1,2})\.(\d{1,2})\.(\d{2})$"),
    ("%d-%m-%Y", r"^(\d{1,2})-(\d{1,2})-(\d{4})$"),
    ("%d-%m-%y", r"^(\d{1,2})-(\d{1,2})-(\d{2})$"),
)
# `d.M` with no year at all (Sample B). The year is not in the file and must
# come from context -- see `_resolve_year`.
_DAY_MONTH = re.compile(r"^(\d{1,2})[./](\d{1,2})$")


def _parse_date(value: Any) -> str:
    """A header cell as an ISO date, or empty.

    Handles both source formats plus real `datetime` values, which is what
    openpyxl hands back when Excel stored the header as a date rather than
    as text. A Hebrew weekday name may trail the date (`2.2 ראשון`) and is
    stripped before parsing.
    """
    if hasattr(value, "isoformat") and not isinstance(value, str):
        return value.isoformat()[:10]
    if (
        isinstance(value, (int, float))
        and not isinstance(value, bool)
        and 20000 <= value <= 80000
    ):
        return (
            datetime.date(1899, 12, 30) + datetime.timedelta(days=int(value))
        ).isoformat()
    text = _normalise(value)
    if not text:
        return ""
    # Strip a trailing or leading weekday name: it is a label on the date,
    # not part of it.
    for weekday in _HEBREW_WEEKDAYS:
        text = text.replace("יום " + weekday, " ").replace(weekday, " ")
    text = text.strip(" ,|-()׳״")
    if not text:
        return ""

    for fmt, pattern in _DATE_PATTERNS:
        if re.match(pattern, text):
            try:
                return datetime.datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
    match = _DAY_MONTH.match(text)
    if match:
        day, month = int(match.group(1)), int(match.group(2))
        return _resolve_year(day, month)
    return ""


def _resolve_year(day: int, month: int) -> str:
    """A `d.M` date with the year the file never wrote.

    Sample B's dates carry no year. Assuming the current one is the only
    defensible default without knowing the period being imported, and it is
    surfaced as a warning on the interpretation rather than applied silently
    — a schedule filed under the wrong year is exactly the kind of quiet
    wrongness the confirm step exists to catch.
    """
    year = datetime.date.today().year
    try:
        return datetime.date(year, month, day).isoformat()
    except ValueError:
        return ""


def _warnings(dates: List[tuple]) -> List[str]:
    """What the manager should look at twice before approving."""
    notes = []
    years = set()
    for _column, iso in dates:
        if len(iso) >= 4:
            years.add(iso[:4])
    if len(years) > 1:
        notes.append("הקובץ מכיל תאריכים מיותר משנה אחת — ודא שהשנה נכונה")
    return notes


# -- helpers ---------------------------------------------------------------

def _vocabulary(profile: Optional[dict]) -> List[dict]:
    """The declared shifts, with their hours (D9).

    Carries `start_time`/`end_time` as well as the name because a real sheet
    often heads its columns with hours rather than names, and the hours are
    what identifies the shift then.
    """
    shifts = (profile or {}).get("shifts") or []
    found: List[dict] = []
    seen = set()
    for shift in shifts:
        if isinstance(shift, dict):
            name = _text(shift.get("name"))
            start = _text(shift.get("start_time"))
            end = _text(shift.get("end_time"))
        else:
            name, start, end = _text(shift), "", ""
        if not name or name in seen:
            continue
        seen.add(name)
        found.append({"name": name, "start_time": start, "end_time": end})
    return found


def _bounded(grid: List[List[str]]) -> List[List[str]]:
    return [row[:_MAX_COLUMNS] for row in (grid or [])[:_MAX_ROWS]]


def _rectangular(grid: List[List[str]]) -> List[List[str]]:
    width = max((len(row) for row in grid), default=0)
    return [list(row) + [""] * (width - len(row)) for row in grid]


def _lane_label(row: List[str], data_columns: List[int]) -> str:
    """The row's own label, taken only from outside the data columns.

    `data_columns` are the cells the grid's header already claimed. A label
    is what sits beside them, so reading the first non-empty cell anywhere
    would return an assignment as though it were the row's name.
    """
    claimed = set(data_columns)
    for column, value in enumerate(row):
        if column in claimed:
            continue
        text = _text(value)
        if text:
            return text
    return ""


def _lane_from_cells(row: List[str], data_columns: List[int]) -> str:
    """Who a label-less lane belongs to, read from the names it holds.

    Sample B writes the person's name into the cell where they are placed
    rather than in a lane column, so the lane's identity *is* its names. An
    availability marker names nobody, which is the whole reason this cannot
    just take the first value in the row.
    """
    for column in data_columns:
        if column >= len(row):
            continue
        for name in _split_names(row[column]):
            if not _is_unavailable(name):
                return name
    return ""


def _mostly_shifts(
    row: List[str], dates: List[tuple], vocabulary: List[dict]
) -> bool:
    """Whether a row's data cells are shift names rather than people.

    Evidence that a row is a second header line: Sample B's nested header
    puts a shift name in every column, and a reader expecting people there
    would otherwise invent one person per shift name.

    Tested against the **declared** names only, never `_match_shift`'s
    fallback. That fallback deliberately accepts any unmatched text so a file
    can name shifts the interview never heard of -- which means it accepts
    people's names too, and using it here would classify every ordinary data
    row as a header and read nothing at all.
    """
    if not vocabulary:
        return False
    filled = matched = 0
    for column, _iso in dates:
        if column >= len(row) or not row[column]:
            continue
        filled += 1
        if _declared_shift(row[column], vocabulary):
            matched += 1
    return filled > 0 and matched == filled


def _declared_shift(value: str, vocabulary: List[dict]) -> str:
    """A header cell matched strictly against the declared vocabulary.

    The strict half of `_match_shift`, without its "take the sheet's own
    word" fallback. Callers that need to know whether a cell is *recognised*
    -- rather than what to call it -- must use this one.
    """
    text = _normalise(value)
    if not text or _parse_date(value) or text in _HEBREW_WEEKDAYS:
        return ""
    for shift in vocabulary:
        declared = _normalise(shift.get("name"))
        if declared and (text == declared or declared in text):
            return shift.get("name")
    hours = _parse_hours(text)
    if hours:
        for shift in vocabulary:
            declared_hours = _parse_hours("%s-%s" % (
                shift.get("start_time") or "", shift.get("end_time") or "",
            ))
            if declared_hours and declared_hours == hours:
                return shift.get("name")
    return ""


def _normalise(value: Any) -> str:
    """Trimmed, whitespace-collapsed text for comparison."""
    text = _text(value).translate({
        ord("\u200e"): None,
        ord("\u200f"): None,
        ord("\ufeff"): None,
        ord("\u00a0"): " ",
    })
    return re.sub(r"\s+", " ", text).strip()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if hasattr(value, "isoformat"):
        return value.isoformat()[:10]
    return str(value).strip()


def _human(iso: str) -> str:
    try:
        day = datetime.datetime.strptime(iso[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return iso
    return "%d.%d.%d" % (day.day, day.month, day.year)


__all__ = ["Interpretation", "read_grid", "infer"]
